#!/usr/bin/env python3
"""
PMO DOMAINS WORKBOOK BUILDER — Generates Excel workbooks for all 12 domains
Style: Ras Flat Dark | Font: Calibri Bold | Headers: #2C3E50 bg, #FFFFFF text
"""
import csv, sqlite3, os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, DataBarRule

BASE = Path(__file__).parent

# ============================================================================
# STYLE CONSTANTS (Ras Flat Dark)
# ============================================================================
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color="2C3E50", size=14)
SUBTITLE_FONT = Font(name="Calibri", bold=True, color="7F8C8D", size=10)
DATA_FONT = Font(name="Calibri", size=10)
LINK_FONT = Font(name="Calibri", size=10, color="2980B9", underline="single")
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

# Domain colors for sheet tabs
TAB_COLORS = {
    "01": "00B4D8", "02": "06D6A0", "03": "FFD166", "04": "EF476F",
    "05": "118AB2", "06": "073B4C", "07": "8338EC", "08": "FF006E",
    "09": "FB5607", "10": "3A86FF", "11": "80ED99", "12": "C77DFF",
}

def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER

def style_data_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER

def auto_width(ws, num_cols, min_width=12, max_width=35):
    for col in range(1, num_cols + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len

def add_title(ws, title, subtitle, row=1):
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    ws.cell(row=row+1, column=1, value=subtitle).font = SUBTITLE_FONT
    return row + 3

def load_csv_data(csv_path):
    with open(csv_path, encoding="utf-8") as f:
        reader = csv.reader(f)
        return list(reader)

def write_table(ws, data, start_row=1, start_col=1):
    if not data:
        return start_row
    for r_idx, row in enumerate(data):
        for c_idx, val in enumerate(row):
            cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx)
            # Try to convert numbers
            try:
                if '.' in val:
                    cell.value = float(val)
                elif val.isdigit():
                    cell.value = int(val)
                else:
                    cell.value = val
            except (ValueError, AttributeError):
                cell.value = val
    # Style header
    style_header_row(ws, start_row, len(data[0]))
    # Style data rows
    for r in range(start_row + 1, start_row + len(data)):
        style_data_row(ws, r, len(data[0]))
    return start_row + len(data) + 1

def add_hyperlinks_column(ws, data, start_row, link_col, file_col, base_dir):
    """Add hyperlinks to a column based on file names in another column"""
    for r in range(start_row + 1, start_row + len(data)):
        cell = ws.cell(row=r, column=link_col)
        file_cell = ws.cell(row=r, column=file_col)
        if file_cell.value and os.path.exists(base_dir / str(file_cell.value)):
            cell.hyperlink = str(base_dir / file_cell.value)
            cell.value = "Open"
            cell.font = LINK_FONT

def add_conditional_formatting(ws, col_letter, start_row, end_row, color_scale="green-yellow-red"):
    """Add conditional formatting to a column"""
    if color_scale == "green-yellow-red":
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"))
        )
        ws.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"))
        )

def add_data_bars(ws, col_letter, start_row, end_row, color="3A86FF"):
    """Add data bars to a column"""
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        DataBarRule(start_type="min", end_type="max", color=color)
    )

# ============================================================================
# WORKBOOK 01: PROJECT & PLOT
# ============================================================================
def build_workbook_01():
    d = BASE / "01_PROJECT_AND_PLOT"
    wb = Workbook()
    wb.properties.title = "01 - Project & Plot Data"
    wb.properties.subject = "Project Identity and Plot Geometry"

    # Sheet 1: Project Master
    ws = wb.active
    ws.title = "Project Master"
    ws.sheet_properties.tabColor = TAB_COLORS["01"]
    row = add_title(ws, "01 — المشروع والأراضي | Project & Plot", "بيانات المشاريع الرئيسية وصفات الأراضي")

    data = load_csv_data(d / "data_raw" / "project_data.csv")
    end_row = write_table(ws, data, start_row=row)
    auto_width(ws, len(data[0]))
    add_data_bars(ws, "O", row + 1, end_row - 1, "00B4D8")  # total_cost
    add_data_bars(ws, "P", row + 1, end_row - 1, "06D6A0")  # cost_per_sqm

    # Sheet 2: Schema Reference
    ws2 = wb.create_sheet("Schema Reference")
    ws2.sheet_properties.tabColor = TAB_COLORS["01"]
    row2 = add_title(ws2, "Schema Reference", "تعريفات الجداول")
    sql_content = (d / "schemas" / "01_project_plot.sql").read_text()
    for i, line in enumerate(sql_content.split("\n")):
        ws2.cell(row=row2 + i, column=1, value=line).font = Font(name="Consolas", size=9, color="2C3E50")
    ws2.column_dimensions["A"].width = 100

    # Sheet 3: Diagrams
    ws3 = wb.create_sheet("Diagrams")
    ws3.sheet_properties.tabColor = TAB_COLORS["01"]
    row3 = add_title(ws3, "ERD Diagram", "مخطط العلاقات")
    ws3.cell(row=row3, column=1, value="Diagram File:").font = BOLD_FONT
    ws3.cell(row=row3, column=2, value="project_plot_erd.mmd").font = DATA_FONT
    ws3.cell(row=row3+1, column=1, value="Database:").font = BOLD_FONT
    ws3.cell(row=row3+1, column=2, value="project_plot.db").font = DATA_FONT
    ws3.cell(row=row3+1, column=2).hyperlink = str(d / "databases" / "project_plot.db")
    ws3.cell(row=row3+1, column=2).font = LINK_FONT

    wb.save(d / "workbook.xlsx")
    print(f"  [XLSX] 01_PROJECT_AND_PLOT/workbook.xlsx")

# ============================================================================
# WORKBOOK 02: LOCATION & GIS
# ============================================================================
def build_workbook_02():
    d = BASE / "02_LOCATION_GIS"
    wb = Workbook()
    wb.properties.title = "02 - Location & GIS"

    ws = wb.active
    ws.title = "UAE Emirates"
    ws.sheet_properties.tabColor = TAB_COLORS["02"]
    row = add_title(ws, "02 — الموقع والخرائط | Location & GIS", "الإمارات والجهات الرقابية والتنقلات الجغرافية")

    data = load_csv_data(d / "data_raw" / "location_data.csv")
    end_row = write_table(ws, data, start_row=row)
    auto_width(ws, len(data[0]))

    # Schema sheet
    ws2 = wb.create_sheet("Schema")
    ws2.sheet_properties.tabColor = TAB_COLORS["02"]
    row2 = add_title(ws2, "Schema Reference", "")
    sql_content = (d / "schemas" / "02_location_gis.sql").read_text()
    for i, line in enumerate(sql_content.split("\n")):
        ws2.cell(row=row2 + i, column=1, value=line).font = Font(name="Consolas", size=9, color="2C3E50")
    ws2.column_dimensions["A"].width = 100

    wb.save(d / "workbook.xlsx")
    print(f"  [XLSX] 02_LOCATION_GIS/workbook.xlsx")

# ============================================================================
# WORKBOOK 03: STAKEHOLDERS
# ============================================================================
def build_workbook_03():
    d = BASE / "03_STAKEHOLDERS"
    wb = Workbook()
    wb.properties.title = "03 - Stakeholders"

    ws = wb.active
    ws.title = "Stakeholders"
    ws.sheet_properties.tabColor = TAB_COLORS["03"]
    row = add_title(ws, "03 — أصحاب المصلحة | Stakeholders", "المالكون والمطورون والمقاولون والاستشاريون")

    data = load_csv_data(d / "data_raw" / "stakeholders_data.csv")
    end_row = write_table(ws, data, start_row=row)
    auto_width(ws, len(data[0]))

    ws2 = wb.create_sheet("Schema")
    ws2.sheet_properties.tabColor = TAB_COLORS["03"]
    row2 = add_title(ws2, "Schema Reference", "")
    sql_content = (d / "schemas" / "03_stakeholders.sql").read_text()
    for i, line in enumerate(sql_content.split("\n")):
        ws2.cell(row=row2 + i, column=1, value=line).font = Font(name="Consolas", size=9, color="2C3E50")
    ws2.column_dimensions["A"].width = 100

    wb.save(d / "workbook.xlsx")
    print(f"  [XLSX] 03_STAKEHOLDERS/workbook.xlsx")

# ============================================================================
# WORKBOOK 04: ZONING & REGULATORY
# ============================================================================
def build_workbook_04():
    d = BASE / "04_ZONING_REGULATORY"
    wb = Workbook()
    wb.properties.title = "04 - Zoning & Regulatory"

    ws = wb.active
    ws.title = "FAR Limits"
    ws.sheet_properties.tabColor = TAB_COLORS["04"]
    row = add_title(ws, "04 — التخطيط والتنظيم | Zoning & Regulatory", "حدود FAR والتغطية والارتفاع حسب الإمارة")

    data = load_csv_data(d / "data_raw" / "far_limits.csv")
    end_row = write_table(ws, data, start_row=row)
    auto_width(ws, len(data[0]))

    # Add data bars for FAR
    add_data_bars(ws, "D", row + 1, end_row - 1, "EF476F")

    ws2 = wb.create_sheet("Schema")
    ws2.sheet_properties.tabColor = TAB_COLORS["04"]
    row2 = add_title(ws2, "Schema Reference", "")
    sql_content = (d / "schemas" / "04_zoning_regulatory.sql").read_text()
    for i, line in enumerate(sql_content.split("\n")):
        ws2.cell(row=row2 + i, column=1, value=line).font = Font(name="Consolas", size=9, color="2C3E50")
    ws2.column_dimensions["A"].width = 100

    wb.save(d / "workbook.xlsx")
    print(f"  [XLSX] 04_ZONING_REGULATORY/workbook.xlsx")

# ============================================================================
# WORKBOOK 05: VALIDATION & COMPLIANCE
# ============================================================================
def build_workbook_05():
    d = BASE / "05_VALIDATION_COMPLIANCE"
    wb = Workbook()
    wb.properties.title = "05 - Validation & Compliance"

    ws = wb.active
    ws.title = "Compliance Checks"
    ws.sheet_properties.tabColor = TAB_COLORS["05"]
    row = add_title(ws, "05 — المطابقة والتوثيق | Validation & Compliance", "فحوصات المطابقة والتصاريح")

    data = load_csv_data(d / "data_raw" / "compliance_data.csv")
    end_row = write_table(ws, data, start_row=row)
    auto_width(ws, len(data[0]))

    ws2 = wb.create_sheet("Schema")
    ws2.sheet_properties.tabColor = TAB_COLORS["05"]
    row2 = add_title(ws2, "Schema Reference", "")
    sql_content = (d / "schemas" / "05_validation.sql").read_text()
    for i, line in enumerate(sql_content.split("\n")):
        ws2.cell(row=row2 + i, column=1, value=line).font = Font(name="Consolas", size=9, color="2C3E50")
    ws2.column_dimensions["A"].width = 100

    wb.save(d / "workbook.xlsx")
    print(f"  [XLSX] 05_VALIDATION_COMPLIANCE/workbook.xlsx")

# ============================================================================
# WORKBOOK 06: DESIGN PARAMETERS
# ============================================================================
def build_workbook_06():
    d = BASE / "06_DESIGN_PARAMETERS"
    wb = Workbook()
    wb.properties.title = "06 - Design Parameters"

    ws = wb.active
    ws.title = "Cost Indices"
    ws.sheet_properties.tabColor = TAB_COLORS["06"]
    row = add_title(ws, "06 — التصميم والمعايير | Design Parameters", "مؤشرات التكلفة حسب الإمارة والنوع")

    data = load_csv_data(d / "data_raw" / "cost_indices.csv")
    end_row = write_table(ws, data, start_row=row)
    auto_width(ws, len(data[0]))
    add_data_bars(ws, "C", row + 1, end_row - 1, "073B4C")

    ws2 = wb.create_sheet("Schema")
    ws2.sheet_properties.tabColor = TAB_COLORS["06"]
    row2 = add_title(ws2, "Schema Reference", "")
    sql_content = (d / "schemas" / "06_design.sql").read_text()
    for i, line in enumerate(sql_content.split("\n")):
        ws2.cell(row=row2 + i, column=1, value=line).font = Font(name="Consolas", size=9, color="2C3E50")
    ws2.column_dimensions["A"].width = 100

    wb.save(d / "workbook.xlsx")
    print(f"  [XLSX] 06_DESIGN_PARAMETERS/workbook.xlsx")

# ============================================================================
# WORKBOOK 07: UNIT MIX & PROGRAM
# ============================================================================
def build_workbook_07():
    d = BASE / "07_UNIT_MIX_PROGRAM"
    wb = Workbook()
    wb.properties.title = "07 - Unit Mix & Program"

    ws = wb.active
    ws.title = "Unit Types"
    ws.sheet_properties.tabColor = TAB_COLORS["07"]
    row = add_title(ws, "07 — الوحدات والبرامج | Unit Mix & Program", "أنواع الوحدات والمساحات")

    data = load_csv_data(d / "data_raw" / "unit_types.csv")
    end_row = write_table(ws, data, start_row=row)
    auto_width(ws, len(data[0]))
    add_data_bars(ws, "E", row + 1, end_row - 1, "8338EC")

    ws2 = wb.create_sheet("Schema")
    ws2.sheet_properties.tabColor = TAB_COLORS["07"]
    row2 = add_title(ws2, "Schema Reference", "")
    sql_content = (d / "schemas" / "07_unit_mix.sql").read_text()
    for i, line in enumerate(sql_content.split("\n")):
        ws2.cell(row=row2 + i, column=1, value=line).font = Font(name="Consolas", size=9, color="2C3E50")
    ws2.column_dimensions["A"].width = 100

    wb.save(d / "workbook.xlsx")
    print(f"  [XLSX] 07_UNIT_MIX_PROGRAM/workbook.xlsx")

# ============================================================================
# WORKBOOK 08: COST & ECONOMICS
# ============================================================================
def build_workbook_08():
    d = BASE / "08_COST_ECONOMICS"
    wb = Workbook()
    wb.properties.title = "08 - Cost & Economics"

    # Sheet 1: CSI Cost Codes
    ws = wb.active
    ws.title = "CSI Cost Codes"
    ws.sheet_properties.tabColor = TAB_COLORS["08"]
    row = add_title(ws, "08 — التكاليف والتحليل المالي | Cost & Economics", "أكواد CSI ومؤشرات التكلفة لكل فئة جودة")

    data = load_csv_data(d / "data_raw" / "csi_cost_codes.csv")
    end_row = write_table(ws, data, start_row=row)
    auto_width(ws, len(data[0]))
    add_data_bars(ws, "D", row + 1, end_row - 1, "FF006E")
    add_data_bars(ws, "E", row + 1, end_row - 1, "FF006E")
    add_data_bars(ws, "F", row + 1, end_row - 1, "FF006E")
    add_data_bars(ws, "G", row + 1, end_row - 1, "FF006E")

    # Sheet 2: Cost Benchmarks
    ws2 = wb.create_sheet("Cost Benchmarks")
    ws2.sheet_properties.tabColor = TAB_COLORS["08"]
    row2 = add_title(ws2, "Cost Benchmarks", "أسعار المراجع لكل بند")
    data2 = load_csv_data(d / "data_raw" / "cost_benchmarks.csv")
    end_row2 = write_table(ws2, data2, start_row=row2)
    auto_width(ws2, len(data2[0]))
    add_data_bars(ws2, "C", row2 + 1, end_row2 - 1, "FF006E")

    # Sheet 3: Lifecycle Costs
    ws3 = wb.create_sheet("Lifecycle Costs")
    ws3.sheet_properties.tabColor = TAB_COLORS["08"]
    row3 = add_title(ws3, "UAE Real Estate Lifecycle Costs", "تكاليف دورة حياة العقار الإماراتي")
    data3 = load_csv_data(d / "data_raw" / "uae_lifecycle_costs.csv")
    end_row3 = write_table(ws3, data3, start_row=row3)
    auto_width(ws3, len(data3[0]))

    # Add total row with formula
    total_row = end_row3
    ws3.cell(row=total_row, column=1, value="TOTAL").font = BOLD_FONT
    ws3.cell(row=total_row, column=3).font = BOLD_FONT
    ws3.cell(row=total_row, column=3, value=f"=SUM(C{row3}:C{end_row3-1})")
    ws3.cell(row=total_row, column=3).number_format = '#,##0'

    # Sheet 4: Schema
    ws4 = wb.create_sheet("Schema")
    ws4.sheet_properties.tabColor = TAB_COLORS["08"]
    row4 = add_title(ws4, "Schema Reference", "")
    sql_content = (d / "schemas" / "08_cost_economics.sql").read_text()
    for i, line in enumerate(sql_content.split("\n")):
        ws4.cell(row=row4 + i, column=1, value=line).font = Font(name="Consolas", size=9, color="2C3E50")
    ws4.column_dimensions["A"].width = 100

    wb.save(d / "workbook.xlsx")
    print(f"  [XLSX] 08_COST_ECONOMICS/workbook.xlsx")

# ============================================================================
# WORKBOOK 09: SCHEDULE & TIMELINE
# ============================================================================
def build_workbook_09():
    d = BASE / "09_SCHEDULE_TIMELINE"
    wb = Workbook()
    wb.properties.title = "09 - Schedule & Timeline"

    ws = wb.active
    ws.title = "WBS Activities"
    ws.sheet_properties.tabColor = TAB_COLORS["09"]
    row = add_title(ws, "09 — الخطة الزمنية | Schedule & Timeline", "أنشطة جدول الأعمال مع المدد الحقيقية")

    data = load_csv_data(d / "data_raw" / "wbs_activities.csv")
    end_row = write_table(ws, data, start_row=row)
    auto_width(ws, len(data[0]))
    add_data_bars(ws, "E", row + 1, end_row - 1, "FB5607")  # duration_days

    # Sheet 2: Summary by Phase
    ws2 = wb.create_sheet("Phase Summary")
    ws2.sheet_properties.tabColor = TAB_COLORS["09"]
    row2 = add_title(ws2, "Phase Summary", "ملخص المراحل")
    ws2.cell(row=row2, column=1, value="Phase").font = HEADER_FONT
    ws2.cell(row=row2, column=1).fill = HEADER_FILL
    ws2.cell(row=row2, column=2, value="Total Duration (days)").font = HEADER_FONT
    ws2.cell(row=row2, column=2).fill = HEADER_FILL
    ws2.cell(row=row2, column=3, value="Activity Count").font = HEADER_FONT
    ws2.cell(row=row2, column=3).fill = HEADER_FILL
    for c in range(1, 4):
        ws2.cell(row=row2, column=c).border = THIN_BORDER
        ws2.cell(row=row2, column=c).alignment = CENTER

    # Calculate phase summaries
    phases = {}
    for r in range(row + 1, end_row):
        phase = ws.cell(row=r, column=2).value
        dur = ws.cell(row=r, column=5).value
        if phase and dur:
            if phase not in phases:
                phases[phase] = {"duration": 0, "count": 0}
            phases[phase]["duration"] += dur
            phases[phase]["count"] += 1

    r2 = row2 + 1
    for phase, info in phases.items():
        ws2.cell(row=r2, column=1, value=phase).font = DATA_FONT
        ws2.cell(row=r2, column=1).border = THIN_BORDER
        ws2.cell(row=r2, column=2, value=info["duration"]).font = DATA_FONT
        ws2.cell(row=r2, column=2).border = THIN_BORDER
        ws2.cell(row=r2, column=2).number_format = '#,##0'
        ws2.cell(row=r2, column=3, value=info["count"]).font = DATA_FONT
        ws2.cell(row=r2, column=3).border = THIN_BORDER
        r2 += 1
    auto_width(ws2, 3)

    ws3 = wb.create_sheet("Schema")
    ws3.sheet_properties.tabColor = TAB_COLORS["09"]
    row3 = add_title(ws3, "Schema Reference", "")
    sql_content = (d / "schemas" / "09_schedule.sql").read_text()
    for i, line in enumerate(sql_content.split("\n")):
        ws3.cell(row=row3 + i, column=1, value=line).font = Font(name="Consolas", size=9, color="2C3E50")
    ws3.column_dimensions["A"].width = 100

    wb.save(d / "workbook.xlsx")
    print(f"  [XLSX] 09_SCHEDULE_TIMELINE/workbook.xlsx")

# ============================================================================
# WORKBOOK 10: APPROVALS & AUTHORITIES
# ============================================================================
def build_workbook_10():
    d = BASE / "10_APPROVALS_AUTHORITIES"
    wb = Workbook()
    wb.properties.title = "10 - Approvals & Authorities"

    ws = wb.active
    ws.title = "Permit Sequence"
    ws.sheet_properties.tabColor = TAB_COLORS["10"]
    row = add_title(ws, "10 — الموافقات والسلطات | Approvals & Authorities", "سلسلة الـ 14 اعتماد مع الأزمنة والتكاليف (min/max/avg)")

    data = load_csv_data(d / "data_raw" / "permit_sequence.csv")
    end_row = write_table(ws, data, start_row=row)
    auto_width(ws, len(data[0]))
    add_data_bars(ws, "E", row + 1, end_row - 1, "3A86FF")  # avg_days
    add_data_bars(ws, "H", row + 1, end_row - 1, "3A86FF")  # avg_cost

    # Add total row
    total_row = end_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = BOLD_FONT
    for col_idx in [5, 8]:  # avg_days, avg_cost
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx).font = BOLD_FONT
        ws.cell(row=total_row, column=col_idx, value=f"=SUM({col_letter}{row+1}:{col_letter}{end_row-1})")
        ws.cell(row=total_row, column=col_idx).number_format = '#,##0'

    ws2 = wb.create_sheet("Schema")
    ws2.sheet_properties.tabColor = TAB_COLORS["10"]
    row2 = add_title(ws2, "Schema Reference", "")
    sql_content = (d / "schemas" / "10_approvals.sql").read_text()
    for i, line in enumerate(sql_content.split("\n")):
        ws2.cell(row=row2 + i, column=1, value=line).font = Font(name="Consolas", size=9, color="2C3E50")
    ws2.column_dimensions["A"].width = 100

    wb.save(d / "workbook.xlsx")
    print(f"  [XLSX] 10_APPROVALS_AUTHORITIES/workbook.xlsx")

# ============================================================================
# WORKBOOK 11: GEOTECHNICAL
# ============================================================================
def build_workbook_11():
    d = BASE / "11_GEOTECHNICAL"
    wb = Workbook()
    wb.properties.title = "11 - Geotechnical"

    # Sheet 1: Soil Types
    ws = wb.active
    ws.title = "Soil Types"
    ws.sheet_properties.tabColor = TAB_COLORS["11"]
    row = add_title(ws, "11 — الجيوتقنية والأساسيات | Geotechnical", "أنواع التربة وخصائصها التحملية والأساسية")

    data = load_csv_data(d / "data_raw" / "soil_types.csv")
    end_row = write_table(ws, data, start_row=row)
    auto_width(ws, len(data[0]))
    add_data_bars(ws, "C", row + 1, end_row - 1, "80ED99")  # bc_typical

    # Sheet 2: Site Specific
    ws2 = wb.create_sheet("Site Data")
    ws2.sheet_properties.tabColor = TAB_COLORS["11"]
    row2 = add_title(ws2, "Site-Specific Geotechnical Data", "بيانات جيوتقنية لكل موقع")
    data2 = load_csv_data(d / "data_raw" / "site_geotechnical.csv")
    end_row2 = write_table(ws2, data2, start_row=row2)
    auto_width(ws2, len(data2[0]))
    add_data_bars(ws2, "F", row2 + 1, end_row2 - 1, "80ED99")  # bearing capacity

    # Sheet 3: Engineering Properties
    ws3 = wb.create_sheet("Engineering Properties")
    ws3.sheet_properties.tabColor = TAB_COLORS["11"]
    row3 = add_title(ws3, "Engineering Soil Properties", "الخصائص الهندسية للتربة (Terzaghi/Meyerhof)")
    data3 = load_csv_data(d / "data_raw" / "engineering_soils.csv")
    end_row3 = write_table(ws3, data3, start_row=row3)
    auto_width(ws3, len(data3[0]))

    # Sheet 4: Schema
    ws4 = wb.create_sheet("Schema")
    ws4.sheet_properties.tabColor = TAB_COLORS["11"]
    row4 = add_title(ws4, "Schema Reference", "")
    sql_content = (d / "schemas" / "11_geotechnical.sql").read_text()
    for i, line in enumerate(sql_content.split("\n")):
        ws4.cell(row=row4 + i, column=1, value=line).font = Font(name="Consolas", size=9, color="2C3E50")
    ws4.column_dimensions["A"].width = 100

    wb.save(d / "workbook.xlsx")
    print(f"  [XLSX] 11_GEOTECHNICAL/workbook.xlsx")

# ============================================================================
# WORKBOOK 12: SUSTAINABILITY & QUALITY & BIM
# ============================================================================
def build_workbook_12():
    d = BASE / "12_SUSTAINABILITY_QUALITY_BIM"
    wb = Workbook()
    wb.properties.title = "12 - Sustainability & BIM"

    ws = wb.active
    ws.title = "Risk Register"
    ws.sheet_properties.tabColor = TAB_COLORS["12"]
    row = add_title(ws, "12 — الاستدامة والجودة/BIM | Sustainability & BIM", "سجل المخاطر مع الاحتمالية والتأثير")

    data = load_csv_data(d / "data_raw" / "risk_catalog.csv")
    end_row = write_table(ws, data, start_row=row)
    auto_width(ws, len(data[0]))
    add_data_bars(ws, "G", row + 1, end_row - 1, "C77DFF")  # risk_score

    ws2 = wb.create_sheet("Schema")
    ws2.sheet_properties.tabColor = TAB_COLORS["12"]
    row2 = add_title(ws2, "Schema Reference", "")
    sql_content = (d / "schemas" / "12_sustainability.sql").read_text()
    for i, line in enumerate(sql_content.split("\n")):
        ws2.cell(row=row2 + i, column=1, value=line).font = Font(name="Consolas", size=9, color="2C3E50")
    ws2.column_dimensions["A"].width = 100

    wb.save(d / "workbook.xlsx")
    print(f"  [XLSX] 12_SUSTAINABILITY_QUALITY_BIM/workbook.xlsx")

# ============================================================================
# MASTER INDEX WORKBOOK
# ============================================================================
def build_master_index():
    wb = Workbook()
    wb.properties.title = "PMO DOMAINS - Master Index"
    wb.properties.subject = "Master Index for all 12 PMO Domains"

    ws = wb.active
    ws.title = "Master Index"
    ws.sheet_properties.tabColor = "2C3E50"

    # Title
    row = 1
    ws.cell(row=row, column=1, value="PMO DOMAINS — الفهرس الرئيسي").font = Font(name="Calibri", bold=True, color="2C3E50", size=18)
    ws.cell(row=row+1, column=1, value="Master Index — 12 Domain Packages").font = SUBTITLE_FONT
    ws.cell(row=row+2, column=1, value="Generated: 2026-06-16 | Style: Ras Flat Dark").font = SUBTITLE_FONT
    row = row + 4

    # Headers
    headers = ["#", "Domain (EN)", "النطاق (AR)", "Schema", "Data Files", "Database", "Color", "Description"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    # Data
    domains = [
        ("01", "Project & Plot", "المشروع والأراضي", "01_project_plot.sql", "project_data.csv", "project_plot.db", "00B4D8", "Project identity and plot geometry"),
        ("02", "Location & GIS", "الموقع والخرائط", "02_location_gis.sql", "location_data.csv", "location_gis.db", "06D6A0", "Emirates jurisdictions and GIS data"),
        ("03", "Stakeholders", "أصحاب المصلحة", "03_stakeholders.sql", "stakeholders_data.csv", "stakeholders.db", "FFD166", "Owner developer contractor registry"),
        ("04", "Zoning & Regulatory", "التخطيط والتنظيم", "04_zoning_regulatory.sql", "far_limits.csv", "zoning_regulatory.db", "EF476F", "FAR height coverage limits"),
        ("05", "Validation & Compliance", "المطابقة والتوثيق", "05_validation.sql", "compliance_data.csv", "validation.db", "118AB2", "Permit compliance and SLA tracking"),
        ("06", "Design Parameters", "التصميم والمعايير", "06_design.sql", "cost_indices.csv", "design_parameters.db", "073B4C", "Cost indices by emirate and typology"),
        ("07", "Unit Mix & Program", "الوحدات والبرامج", "07_unit_mix.sql", "unit_types.csv", "unit_mix.db", "8338EC", "Unit types and areas"),
        ("08", "Cost & Economics", "التكاليف والتحليل المالي", "08_cost_economics.sql", "csi_cost_codes.csv", "cost_economics.db", "FF006E", "CSI codes benchmarks lifecycle"),
        ("09", "Schedule & Timeline", "الخطة الزمنية", "09_schedule.sql", "wbs_activities.csv", "schedule.db", "FB5607", "WBS activities durations"),
        ("10", "Approvals & Authorities", "الموافقات والسلطات", "10_approvals.sql", "permit_sequence.csv", "approvals.db", "3A86FF", "14-step permit sequence"),
        ("11", "Geotechnical", "الجيوقيقة والأساسيات", "11_geotechnical.sql", "soil_types.csv", "geotechnical.db", "80ED99", "Soil types bearing capacity"),
        ("12", "Sustainability & BIM", "الاستدامة والجودة", "12_sustainability.sql", "risk_catalog.csv", "sustainability.db", "C77DFF", "LEED Estidama quality risk"),
    ]

    for r_idx, domain in enumerate(domains):
        r = row + 1 + r_idx
        for c_idx, val in enumerate(domain):
            cell = ws.cell(row=r, column=c_idx + 1, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER if c_idx < 7 else LEFT
            # Color code the color column
            if c_idx == 6:
                cell.fill = PatternFill(start_color=val, end_color=val, fill_type="solid")
                cell.font = Font(name="Calibri", size=10, color="FFFFFF" if val in ["2C3E50","073B4C","FF006E","EF476F"] else "000000")

    # Column widths
    widths = [5, 22, 20, 25, 22, 25, 10, 40]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    # Summary row
    summary_row = row + len(domains) + 3
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = BOLD_FONT
    ws.cell(row=summary_row, column=2, value="12 Domains").font = BOLD_FONT
    ws.cell(row=summary_row, column=3, value="12 Schemas").font = BOLD_FONT
    ws.cell(row=summary_row, column=4, value="17 CSV Files").font = BOLD_FONT
    ws.cell(row=summary_row, column=5, value="13 Databases").font = BOLD_FONT
    ws.cell(row=summary_row, column=6, value="12 Diagrams").font = BOLD_FONT

    # Add chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Files per Domain"
    chart.y_axis.title = "Count"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    # Data for chart (use a helper column)
    chart_data_row = summary_row + 3
    ws.cell(row=chart_data_row, column=1, value="Domain").font = BOLD_FONT
    ws.cell(row=chart_data_row, column=2, value="Files").font = BOLD_FONT
    file_counts = [4, 3, 3, 3, 3, 3, 3, 5, 4, 3, 5, 3]  # files per domain
    for i, (domain, count) in enumerate(zip([d[1] for d in domains], file_counts)):
        ws.cell(row=chart_data_row + 1 + i, column=1, value=domain).font = DATA_FONT
        ws.cell(row=chart_data_row + 1 + i, column=2, value=count).font = DATA_FONT

    data_ref = Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_row + 12)
    cats_ref = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + 12)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "3A86FF"
    ws.add_chart(chart, f"A{chart_data_row + 15}")

    wb.save(BASE / "MASTER_INDEX.xlsx")
    print(f"  [XLSX] MASTER_INDEX.xlsx")

# ============================================================================
# MAIN
# ============================================================================
if __name__ == "__main__":
    print("=" * 60)
    print("PMO DOMAINS WORKBOOK BUILDER — 12 Excel Workbooks + Master Index")
    print("Style: Ras Flat Dark | Font: Calibri Bold | Headers: #2C3E50")
    print("=" * 60)
    build_workbook_01()
    build_workbook_02()
    build_workbook_03()
    build_workbook_04()
    build_workbook_05()
    build_workbook_06()
    build_workbook_07()
    build_workbook_08()
    build_workbook_09()
    build_workbook_10()
    build_workbook_11()
    build_workbook_12()
    build_master_index()
    print("=" * 60)
    print("ALL WORKBOOKS BUILT SUCCESSFULLY")
    print("=" * 60)
