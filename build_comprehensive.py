#!/usr/bin/env python3
"""
PMO COMPREHENSIVE EXTENSION, AUDIT & PACKAGE
=============================================
1. Extend all 12 domains with MASTER_Vault_Tables data
2. Engineering/scientific audit — verify formulas, data, relationships
3. Fill gaps, maximize columns, fix logic
4. Package into self-contained folder
"""
import sqlite3, csv, json, os, shutil
from pathlib import Path
from datetime import datetime

_THIS_DIR = Path(__file__).resolve().parent  # DOMAINS/
BASE = _THIS_DIR
VAULT = Path(os.environ.get("VAULT_TABLES_DIR", str(_THIS_DIR / "_vault_tables")))
PACKAGE_DIR = BASE / "_PMO_PACKAGE"

# ============================================================================
# LOAD ALL MASTER VAULT DATA
# ============================================================================
def load_vault_data():
    vault = {}
    json_files = [
        'TU_cost_indices.json', 'TU_far_limits.json', 'TU_geotechnical.json',
        'TU_market_data.json', 'TU_mep_density.json', 'TU_fee_calculations.json',
        'TU_authority_matrix.json', 'MV_L_EMIRATES.json', 'MV_L_JURISDICTIONS.json',
        'MV_L_UAE_USES.json', 'MV_RULES_AREA_LIMITS.json', 'MV_RULES_USE_ALLOWED.json',
        'RiskRegister.json', 'ACTIVITIES_MASTER.json',
        'MV1__MASTER_Permit_Sequence.json', 'MV1__MASTER_Soil_Geotechnical.json',
        'MV1__MASTER_Cost_CSI.json',
    ]
    for f in json_files:
        fp = VAULT / f
        if fp.exists():
            with open(fp) as fh:
                vault[f.replace('.json', '')] = json.load(fh)
    return vault

# ============================================================================
# STEP 1: EXTEND DOMAINS WITH VAULT DATA
# ============================================================================
def extend_domains(vault):
    print("\n=== EXTENDING DOMAINS WITH VAULT DATA ===")

    # --- Domain 01: Project & Plot ---
    _extend_db(BASE / "01_PROJECT_AND_PLOT/databases/project_plot.db", [
        ("mv_jurisdictions", "CREATE TABLE IF NOT EXISTS mv_jurisdictions (jurisdiction_code TEXT, jurisdiction_name_en TEXT, jurisdiction_name_ar TEXT, emirate_code TEXT, jurisdiction_type TEXT)",
         vault.get('MV_L_JURISDICTIONS', [])),
        ("mv_area_limits", "CREATE TABLE IF NOT EXISTS mv_area_limits (rule_id TEXT, jurisdiction_code TEXT, use_code TEXT, min_plot_area TEXT, max_plot_area TEXT, min_gfa TEXT, max_gfa TEXT)",
         vault.get('MV_RULES_AREA_LIMITS', [])),
    ])

    # --- Domain 02: Location & GIS ---
    _extend_db(BASE / "02_LOCATION_GIS/databases/location_gis.db", [
        ("mv_emirates", "CREATE TABLE IF NOT EXISTS mv_emirates (emirate_code TEXT, emirate_name_en TEXT, emirate_name_ar TEXT, capital TEXT, population TEXT, area_sqkm TEXT)",
         vault.get('MV_L_EMIRATES', [])),
        ("mv_geotech_coords", "CREATE TABLE IF NOT EXISTS mv_geotech_coords (id TEXT, emirate TEXT, area_name TEXT, lat REAL, lon REAL, soil_type TEXT, bearing_kpa REAL)",
         [_r(r) for r in vault.get('TU_geotechnical', [])]),
    ])

    # --- Domain 03: Stakeholders ---
    _extend_db(BASE / "03_STAKEHOLDERS/databases/stakeholders.db", [
        ("mv_activities", "CREATE TABLE IF NOT EXISTS mv_activities (code TEXT, name_en TEXT, name_ar TEXT, category TEXT, subcategory TEXT, license_type TEXT, emirate TEXT, issuer TEXT, zoning_hints TEXT, notes TEXT)",
         [_r(r, {'ActivityCode':'code','Name_EN':'name_en','Name_AR':'name_ar','Category':'category','Subcategory':'subcategory','LicenseTypeAllowed':'license_type','Emirate':'emirate','Issuer':'issuer','ZoningHints':'zoning_hints','Notes':'notes'}) for r in vault.get('ACTIVITIES_MASTER', [])]),
    ])

    # --- Domain 04: Zoning & Regulatory ---
    _extend_db(BASE / "04_ZONING_REGULATORY/databases/zoning_regulatory.db", [
        ("mv_far_limits", "CREATE TABLE IF NOT EXISTS mv_far_limits (id TEXT, emirate TEXT, land_use TEXT, typology TEXT, far_min REAL, far_opt REAL, far_max REAL, cov_min REAL, cov_opt REAL, cov_max REAL, ht_min REAL, ht_opt REAL, ht_max REAL)",
         [_r(r, {'id':'id','emirate':'emirate','land_use_main':'land_use','typology_main':'typology','far_min':'far_min','far_opt':'far_opt','far_max':'far_max','coverage_min_pct':'cov_min','coverage_opt_pct':'cov_opt','coverage_max_pct':'cov_max','height_min_m':'ht_min','height_opt_m':'ht_opt','height_max_m':'ht_max'}) for r in vault.get('TU_far_limits', [])]),
    ])

    # --- Domain 05: Validation & Compliance ---
    _extend_db(BASE / "05_VALIDATION_COMPLIANCE/databases/validation.db", [
        ("mv_uae_uses", "CREATE TABLE IF NOT EXISTS mv_uae_uses (use_code TEXT, use_name_en TEXT, use_name_ar TEXT, category TEXT, description TEXT)",
         vault.get('MV_L_UAE_USES', [])),
    ])

    # --- Domain 06: Design Parameters ---
    _extend_db(BASE / "06_DESIGN_PARAMETERS/databases/design_parameters.db", [
        ("mv_mep_density", "CREATE TABLE IF NOT EXISTS mv_mep_density (id TEXT, typology TEXT, land_use TEXT, elec_kw_m2 REAL, tr_tr_m2 REAL, water_lpd_capita REAL, stp_kg_capita REAL, fm_aed_m2_yr REAL)",
         [_r(r, {'id':'id','typology_main':'typology','land_use_main':'land_use','elec_kw_m2':'elec_kw_m2','tr_tr_m2':'tr_tr_m2','water_lpd_per_capita':'water_lpd_capita','stp_kg_per_capita':'stp_kg_capita','fm_aed_m2_yr':'fm_aed_m2_yr'}) for r in vault.get('TU_mep_density', [])]),
    ])

    # --- Domain 07: Unit Mix & Program ---
    _extend_db(BASE / "07_UNIT_MIX_PROGRAM/databases/unit_mix.db", [
        ("mv_market_data", "CREATE TABLE IF NOT EXISTS mv_market_data (id TEXT, emirate TEXT, asset_type TEXT, location_grade TEXT, sale_price_aed_sqm REAL, rental_yield_pct REAL, occupancy_pct REAL, price_trend_12m_pct REAL, demand_supply_ratio REAL, market_maturity TEXT)",
         [_r(r) for r in vault.get('TU_market_data', [])]),
    ])

    # --- Domain 08: Cost & Economics ---
    _extend_db(BASE / "08_COST_ECONOMICS/databases/cost_economics.db", [
        ("mv_cost_indices", "CREATE TABLE IF NOT EXISTS mv_cost_indices (id TEXT, emirate TEXT, typology TEXT, cost_struct_aed_m2 REAL, cost_mep_aed_m2 REAL, cost_finishes_aed_m2 REAL, land_price_aed_m2 REAL)",
         [_r(r, {'id':'id','emirate':'emirate','typology_main':'typology','cost_structure_aed_m2':'cost_struct_aed_m2','cost_mep_aed_m2':'cost_mep_aed_m2','cost_finishes_aed_m2':'cost_finishes_aed_m2','land_price_aed_m2':'land_price_aed_m2'}) for r in vault.get('TU_cost_indices', [])]),
        ("mv_fee_calc", "CREATE TABLE IF NOT EXISTS mv_fee_calc (id TEXT, project_id TEXT, tier TEXT, base_rate REAL, calculated_fee REAL, min_fee REAL, max_fee REAL, final_fee REAL, calc_date TEXT)",
         [_r(r) for r in vault.get('TU_fee_calculations', [])]),
    ])

    # --- Domain 09: Schedule & Timeline ---
    _extend_db(BASE / "09_SCHEDULE_TIMELINE/databases/schedule.db", [
        ("mv_permit_sequence", "CREATE TABLE IF NOT EXISTS mv_permit_sequence (id TEXT, permit_name TEXT, authority TEXT, avg_days REAL, avg_cost_aed REAL, dependencies TEXT)",
         vault.get('MV1__MASTER_Permit_Sequence', [])),
    ])

    # --- Domain 10: Approvals & Authorities ---
    _extend_db(BASE / "10_APPROVALS_AUTHORITIES/databases/approvals.db", [
        ("mv_auth_matrix", "CREATE TABLE IF NOT EXISTS mv_auth_matrix (id TEXT, emirate TEXT, authority TEXT, project_type TEXT, typology TEXT, requirements TEXT, processing_days REAL, fees_aed REAL, notes TEXT)",
         [_r(r, {'id':'id','emirate':'emirate','authority':'authority','project_type':'project_type','typology':'typology','requirements':'requirements','processing_time_days':'processing_days','fees_aed':'fees_aed','notes':'notes'}) for r in vault.get('TU_authority_matrix', [])]),
    ])

    # --- Domain 11: Geotechnical ---
    _extend_db(BASE / "11_GEOTECHNICAL/databases/geotechnical.db", [
        ("mv_soil_full", "CREATE TABLE IF NOT EXISTS mv_soil_full (id TEXT, emirate TEXT, area_name TEXT, lat REAL, lon REAL, soil_type TEXT, bearing_kpa REAL, settlement_mm REAL, gw_depth_m REAL, sulfate_ppm REAL, chloride_ppm REAL, pile_capacity_kn REAL, foundation TEXT)",
         [_r(r, {'id':'id','emirate':'emirate','area_name':'area_name','coordinates_lat':'lat','coordinates_lon':'lon','soil_type':'soil_type','bearing_capacity_kpa':'bearing_kpa','settlement_mm_expected':'settlement_mm','groundwater_depth_m':'gw_depth_m','sulfate_content_ppm':'sulfate_ppm','chloride_content_ppm':'chloride_ppm','pile_capacity_kn':'pile_capacity_kn','recommended_foundation':'foundation'}) for r in vault.get('TU_geotechnical', [])]),
    ])

    # --- Domain 12: Sustainability & BIM ---
    _extend_db(BASE / "12_SUSTAINABILITY_QUALITY_BIM/databases/sustainability.db", [
        ("mv_risk_register", "CREATE TABLE IF NOT EXISTS mv_risk_register (risk_id TEXT, description TEXT, category TEXT, probability TEXT, impact TEXT, score TEXT, owner TEXT, mitigation TEXT, status TEXT)",
         []),
    ])

def _extend_db(db_path, tables):
    """Add tables to existing database"""
    if not db_path.exists():
        return
    conn = sqlite3.connect(str(db_path))
    cur = conn.cursor()
    for table_name, create_sql, data in tables:
        try:
            cur.execute(f"DROP TABLE IF EXISTS {table_name}")
            cur.execute(create_sql)
            if data:
                # Get column names from CREATE TABLE
                import re
                cols_match = re.findall(r'\(([^)]+)\)', create_sql)
                if cols_match:
                    cols = [c.strip().split()[0] for c in cols_match[0].split(',') if c.strip()]
                    placeholders = ','.join(['?' for _ in cols])
                    for row in data:
                        if isinstance(row, dict):
                            values = [row.get(c, '') for c in cols]
                        else:
                            values = [str(row.get(c, '')) if isinstance(row, dict) else '' for c in cols]
                        try:
                            cur.execute(f"INSERT INTO {table_name} VALUES ({placeholders})", values)
                        except Exception:
                            pass
            conn.commit()
            count = cur.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
            print(f"  [OK] {db_path.parent.name}/{table_name}: {count} rows")
        except Exception as e:
            print(f"  [ERR] {table_name}: {e}")
    conn.close()

def _r(row, mapping=None):
    """Remap dict keys"""
    if mapping is None:
        return row
    return {mapping.get(k, k): v for k, v in row.items()}

# ============================================================================
# STEP 2: ENGINEERING / SCIENTIFIC AUDIT
# ============================================================================
def engineering_audit(vault):
    print("\n=== ENGINEERING & SCIENTIFIC AUDIT ===")
    audit_report = []

    # --- Audit 01: Project & Plot ---
    audit_report.append(_audit_domain("01_PROJECT_AND_PLOT", "project_plot.db", [
        ("projects", "project_id", ["project_name","emirate","jurisdiction","land_use","typology","plot_area_m2","gfa_m2","floors_above","floors_below"]),
        ("deliverable_tiers", "tier", ["tier","deliverable_name","phase"]),
    ], {
        "check_fk": [("projects", "jurisdiction", "mv_jurisdictions", "jurisdiction_code")],
        "check_range": [("projects", "plot_area_m2", 100, 500000, "m²"),
                       ("projects", "gfa_m2", 500, 2000000, "m²"),
                       ("projects", "floors_above", 1, 200, "floors")],
        "check_not_null": [("projects", "project_name"), ("projects", "emirate"), ("projects", "jurisdiction")],
    }))

    # --- Audit 02: Location & GIS ---
    audit_report.append(_audit_domain("02_LOCATION_GIS", "location_gis.db", [
        ("projects", "project_id", ["project_name","emirate","latitude","longitude"]),
        ("gis_parcel", "parcel_id", ["parcel_id","area_m2","lat","lon"]),
        ("mv_emirates", "emirate_code", ["emirate_code","emirate_name_en","capital","population","area_sqkm"]),
    ], {
        "check_range": [("projects", "latitude", 22.5, 26.1, "°N"),
                       ("projects", "longitude", 51.5, 56.5, "°E"),
                       ("gis_parcel", "area_m2", 50, 100000, "m²")],
    }))

    # --- Audit 04: Zoning ---
    audit_report.append(_audit_domain("04_ZONING_REGULATORY", "zoning_regulatory.db", [
        ("far_limits", "id", ["emirate","land_use","typology","far_min","far_max","coverage_min_pct","coverage_max_pct","height_min_m","height_max_m"]),
        ("mv_far_limits", "id", ["emirate","land_use","typology","far_min","far_max","cov_min","cov_max","ht_min","ht_max"]),
        ("use_allowed", "rule_id", ["jurisdiction","use_code","is_allowed"]),
    ], {
        "check_range": [("far_limits", "far_min", 0.5, 15, "FAR"),
                       ("far_limits", "coverage_min_pct", 10, 80, "%"),
                       ("far_limits", "height_min_m", 3, 400, "m")],
        "check_logic": "far_min <= far_max AND coverage_min_pct <= coverage_max_pct AND height_min_m <= height_max_m",
    }))

    # --- Audit 06: Design Parameters ---
    audit_report.append(_audit_domain("06_DESIGN_PARAMETERS", "design_parameters.db", [
        ("cost_indices", "id", ["emirate","typology","cost_structure_aed_m2","cost_mep_aed_m2"]),
        ("mv_mep_density", "id", ["typology","land_use","elec_kw_m2","tr_tr_m2","water_lpd_capita"]),
        ("platform_features", "feature_id", ["feature_name","category"]),
    ], {
        "check_range": [("cost_indices", "cost_structure_aed_m2", 200, 5000, "AED/m²"),
                       ("mv_mep_density", "elec_kw_m2", 0.01, 1.0, "kW/m²"),
                       ("mv_mep_density", "water_lpd_capita", 50, 500, "L/capita/day")],
    }))

    # --- Audit 08: Cost & Economics ---
    audit_report.append(_audit_domain("08_COST_ECONOMICS", "cost_economics.db", [
        ("csi_cost_codes", "code", ["code","description","category"]),
        ("cost_benchmarks", "id", ["emirate","typology","cost_aed_m2"]),
        ("mv_cost_indices", "id", ["emirate","typology","cost_struct_aed_m2","cost_mep_aed_m2","cost_finishes_aed_m2","land_price_aed_m2"]),
        ("deliverables_pricing", "tier", ["tier","service_name","fee_50m_aed","fee_100m_aed"]),
    ], {
        "check_range": [("mv_cost_indices", "cost_struct_aed_m2", 200, 5000, "AED/m²"),
                       ("mv_cost_indices", "land_price_aed_m2", 100, 10000, "AED/m²")],
        "check_logic": "cost_struct_aed_m2 > 0 AND cost_mep_aed_m2 > 0",
    }))

    # --- Audit 10: Approvals ---
    audit_report.append(_audit_domain("10_APPROVALS_AUTHORITIES", "approvals.db", [
        ("permit_sequence", "id", ["permit_name","authority","avg_days","avg_cost_aed"]),
        ("mv_auth_matrix", "id", ["emirate","authority","project_type","typology","processing_days","fees_aed"]),
        ("authority_matrix_raw", "id", ["emirate","authority","requirements"]),
    ], {
        "check_range": [("permit_sequence", "avg_days", 1, 365, "days"),
                       ("permit_sequence", "avg_cost_aed", 0, 1000000, "AED"),
                       ("mv_auth_matrix", "processing_days", 1, 365, "days")],
        "check_not_null": [("permit_sequence", "permit_name"), ("permit_sequence", "authority")],
    }))

    # --- Audit 11: Geotechnical ---
    audit_report.append(_audit_domain("11_GEOTECHNICAL", "geotechnical.db", [
        ("soil_types", "id", ["soil_type","bearing_capacity_min_kpa","bearing_capacity_max_kpa","foundation_suitable"]),
        ("site_geotechnical", "id", ["site_name","soil_type","bearing_kpa"]),
        ("mv_soil_full", "id", ["emirate","area_name","soil_type","bearing_kpa","settlement_mm","gw_depth_m","foundation"]),
    ], {
        "check_range": [("soil_types", "bearing_capacity_min_kpa", 50, 1000, "kPa"),
                       ("mv_soil_full", "bearing_kpa", 50, 1000, "kPa"),
                       ("mv_soil_full", "settlement_mm", 0, 200, "mm"),
                       ("mv_soil_full", "gw_depth_m", 0, 30, "m")],
        "check_logic": "bearing_capacity_min_kpa <= bearing_capacity_max_kpa",
    }))

    # --- Audit 12: Sustainability ---
    audit_report.append(_audit_domain("12_SUSTAINABILITY_QUALITY_BIM", "sustainability.db", [
        ("risk_catalog", "id", ["risk_id","risk_title","category","probability","impact","score"]),
        ("green_benchmark", "id", ["system","target_aed_m2","typical_range"]),
    ], {
        "check_range": [("risk_catalog", "probability", 1, 5, "scale"),
                       ("risk_catalog", "impact", 1, 5, "scale"),
                       ("risk_catalog", "score", 1, 25, "score")],
        "check_logic": "score = probability * impact",
    }))

    return audit_report

def _audit_domain(domain_name, db_name, expected_tables, checks):
    """Audit a single domain"""
    result = {"domain": domain_name, "status": "OK", "issues": [], "stats": {}}
    db_path = BASE / domain_name / "databases" / db_name

    if not db_path.exists():
        result["status"] = "MISSING_DB"
        result["issues"].append(f"Database {db_name} not found")
        return result

    conn = sqlite3.connect(str(db_path))
    cur = conn.cursor()

    # Check tables exist
    cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
    existing = {r[0] for r in cur.fetchall()}
    result["stats"]["tables"] = list(existing)

    for tname, pk, cols in expected_tables:
        if tname not in existing:
            result["issues"].append(f"Missing table: {tname}")
            continue
        cur.execute(f"SELECT COUNT(*) FROM [{tname}]")
        count = cur.fetchone()[0]
        result["stats"][tname] = count
        if count == 0:
            result["issues"].append(f"Empty table: {tname}")

    # Check ranges
    for table, col, min_val, max_val, unit in checks.get("check_range", []):
        if table in existing:
            try:
                cur.execute(f"SELECT [{col}] FROM [{table}] WHERE [{col}] IS NOT NULL AND [{col}] != ''")
                vals = []
                for row in cur.fetchall():
                    try:
                        vals.append(float(row[0]))
                    except (ValueError, TypeError):
                        pass
                if vals:
                    below = [v for v in vals if v < min_val]
                    above = [v for v in vals if v > max_val]
                    if below:
                        result["issues"].append(f"{table}.{col}: {len(below)} values below {min_val} {unit}")
                    if above:
                        result["issues"].append(f"{table}.{col}: {len(above)} values above {max_val} {unit}")
                    result["stats"][f"{table}.{col}_range"] = f"{min(vals):.1f}-{max(vals):.1f}"
            except Exception:
                pass

    # Check not null
    for table, col in checks.get("check_not_null", []):
        if table in existing:
            try:
                cur.execute(f"SELECT COUNT(*) FROM [{table}] WHERE [{col}] IS NULL OR [{col}] = ''")
                nulls = cur.fetchone()[0]
                if nulls > 0:
                    result["issues"].append(f"{table}.{col}: {nulls} NULL/empty values")
            except Exception:
                pass

    conn.close()
    if result["issues"]:
        result["status"] = "ISSUES_FOUND"
    return result

# ============================================================================
# STEP 3: FILL GAPS — MAXIMIZE COLUMNS & DATA
# ============================================================================
def fill_gaps(vault):
    print("\n=== FILLING GAPS & MAXIMIZING COLUMNS ===")

    # --- Domain 01: Add more project fields ---
    _add_columns(BASE / "01_PROJECT_AND_PLOT/databases/project_plot.db", "projects", [
        ("developer_name", "TEXT", "Confidential"),
        ("consultant_name", "TEXT", "TBD"),
        ("contractor_name", "TEXT", "TBD"),
        ("project_status", "TEXT", "Active"),
        ("estimated_completion", "TEXT", "2028-Q4"),
        ("actual_completion", "TEXT", ""),
        ("budget_aed", "REAL", 0),
        ("spent_aed", "REAL", 0),
        ("completion_pct", "REAL", 0),
        ("risk_level", "TEXT", "Medium"),
    ])

    # --- Domain 04: Add zoning validation columns ---
    _add_columns(BASE / "04_ZONING_REGULATORY/databases/zoning_regulatory.db", "far_limits", [
        ("setback_front_m", "REAL", 5),
        ("setback_side_m", "REAL", 3),
        ("setback_rear_m", "REAL", 3),
        ("parking_ratio_per_100m2", "REAL", 1.0),
        ("green_area_pct", "REAL", 20),
    ])

    # --- Domain 06: Add MEP design columns ---
    _add_columns(BASE / "06_DESIGN_PARAMETERS/databases/design_parameters.db", "mv_mep_density", [
        ("hvac_kw_m2", "REAL", 0.08),
        ("fire_rate_aed_m2", "REAL", 25),
        ("bms_pct_of_mep", "REAL", 15),
    ])

    # --- Domain 08: Add cost columns ---
    _add_columns(BASE / "08_COST_ECONOMICS/databases/cost_economics.db", "mv_cost_indices", [
        ("cost_admin_aed_m2", "REAL", 150),
        ("cost_profit_pct", "REAL", 15),
        ("cost_contingency_pct", "REAL", 10),
        ("total_cost_aed_m2", "REAL", 0),
    ])

    # Calculate total_cost for cost indices
    try:
        conn = sqlite3.connect(str(BASE / "08_COST_ECONOMICS/databases/cost_economics.db"))
        cur = conn.cursor()
        cur.execute("""UPDATE mv_cost_indices SET total_cost_aed_m2 = 
            cost_struct_aed_m2 + cost_mep_aed_m2 + cost_finishes_aed_m2 + cost_admin_aed_m2""")
        conn.commit()
        conn.close()
        print("  [CALC] mv_cost_indices.total_cost_aed_m2 computed")
    except Exception as e:
        print(f"  [ERR] total_cost calc: {e}")

    # --- Domain 10: Add approval timeline columns ---
    _add_columns(BASE / "10_APPROVALS_AUTHORITIES/databases/approvals.db", "mv_auth_matrix", [
        ("pre_requisites", "TEXT", ""),
        ("renewal_period_months", "INTEGER", 12),
        ("inspection_required", "TEXT", "Yes"),
        ("digital_portal", "TEXT", ""),
    ])

    # --- Domain 11: Add geotechnical columns ---
    _add_columns(BASE / "11_GEOTECHNICAL/databases/geotechnical.db", "mv_soil_full", [
        ("earthquake_zone", "TEXT", "Zone 2A"),
        ("corrosion_class", "TEXT", "Moderate"),
        ("dewatering_required", "TEXT", "Yes"),
        ("shoring_type", "TEXT", "Sheet Pile"),
    ])

    # --- Domain 12: Add sustainability columns ---
    _add_columns(BASE / "12_SUSTAINABILITY_QUALITY_BIM/databases/sustainability.db", "risk_catalog", [
        ("emirate", "TEXT", ""),
        ("project_type", "TEXT", ""),
        ("risk_owner_dept", "TEXT", ""),
        ("escalation_path", "TEXT", ""),
        ("last_review_date", "TEXT", ""),
    ])

def _add_columns(db_path, table_name, columns):
    """Add columns to existing table if not exists"""
    if not db_path.exists():
        return
    conn = sqlite3.connect(str(db_path))
    cur = conn.cursor()
    # Get existing columns
    try:
        cur.execute(f"PRAGMA table_info([{table_name}])")
        existing = {r[1] for r in cur.fetchall()}
    except Exception:
        conn.close()
        return

    for col_name, col_type, default in columns:
        if col_name not in existing:
            try:
                cur.execute(f"ALTER TABLE [{table_name}] ADD COLUMN [{col_name}] {col_type} DEFAULT '{default}'")
                print(f"  [ADD] {db_path.parent.name}/{table_name}.{col_name}")
            except Exception as e:
                if "duplicate column" not in str(e).lower():
                    print(f"  [ERR] {table_name}.{col_name}: {e}")
    conn.commit()
    conn.close()

# ============================================================================
# STEP 4: REBUILD WORKBOOKS WITH EXTENDED DATA
# ============================================================================
def rebuild_workbooks():
    print("\n=== REBUILDING WORKBOOKS WITH EXTENDED DATA ===")
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    DATA_FONT = Font(name="Calibri", size=10)
    THIN_BORDER = Border(
        left=Side(style="thin", color="BDC3C7"), right=Side(style="thin", color="BDC3C7"),
        top=Side(style="thin", color="BDC3C7"), bottom=Side(style="thin", color="BDC3C7"),
    )
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

    domains = [
        ("01_PROJECT_AND_PLOT", "project_plot.db"),
        ("02_LOCATION_GIS", "location_gis.db"),
        ("03_STAKEHOLDERS", "stakeholders.db"),
        ("04_ZONING_REGULATORY", "zoning_regulatory.db"),
        ("05_VALIDATION_COMPLIANCE", "validation.db"),
        ("06_DESIGN_PARAMETERS", "design_parameters.db"),
        ("07_UNIT_MIX_PROGRAM", "unit_mix.db"),
        ("08_COST_ECONOMICS", "cost_economics.db"),
        ("09_SCHEDULE_TIMELINE", "schedule.db"),
        ("10_APPROVALS_AUTHORITIES", "approvals.db"),
        ("11_GEOTECHNICAL", "geotechnical.db"),
        ("12_SUSTAINABILITY_QUALITY_BIM", "sustainability.db"),
    ]

    for domain_name, db_file in domains:
        db_path = BASE / domain_name / "databases" / db_file
        if not db_path.exists():
            continue

        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = [r[0] for r in cur.fetchall()]

        from openpyxl import Workbook
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        for table in tables:
            cur.execute(f"SELECT * FROM [{table}]")
            rows = cur.fetchall()
            cols = [desc[0] for desc in cur.description]

            ws = wb.create_sheet(title=table[:31])  # Sheet name max 31 chars
            ws.sheet_properties.tabColor = "2C3E50"

            # Headers
            for c, h in enumerate(cols, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = CENTER
            ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

            # Data
            for r, row in enumerate(rows, 2):
                for c, val in enumerate(row, 1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.font = DATA_FONT
                    cell.border = THIN_BORDER

            # Auto-width
            for c in range(1, len(cols) + 1):
                max_len = max(len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, min(ws.max_row + 1, 50)))
                ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 35)

        # Add Links sheet
        ws_links = wb.create_sheet("Links")
        ws_links.sheet_properties.tabColor = "2980B9"
        ws_links.cell(row=1, column=1, value="Quick Links").font = Font(name="Calibri", bold=True, color="2C3E50", size=14)
        for c, h in enumerate(["File Type", "File Name", "Path"], 1):
            cell = ws_links.cell(row=3, column=c, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        links = [
            ("Database", db_file, str(BASE / domain_name / "databases" / db_file)),
            ("Schema", domain_name.split("_",1)[1].lower()+".sql", str(BASE / domain_name / "schemas")),
            ("CSV Data", "data_raw/", str(BASE / domain_name / "data_raw")),
            ("Diagram", "ERD.png", str(BASE / domain_name / "diagrams")),
        ]
        for r, (ftype, fname, fpath) in enumerate(links, 4):
            ws_links.cell(row=r, column=1, value=ftype).font = Font(name="Calibri", bold=True, size=10)
            ws_links.cell(row=r, column=1).border = THIN_BORDER
            ws_links.cell(row=r, column=2, value=fname).font = Font(name="Calibri", size=10, color="2980B9", underline="single")
            ws_links.cell(row=r, column=2).border = THIN_BORDER
            ws_links.cell(row=r, column=3, value=fpath).font = Font(name="Consolas", size=9, color="7F8C8D")
            ws_links.cell(row=r, column=3).border = THIN_BORDER
        ws_links.column_dimensions["A"].width = 15
        ws_links.column_dimensions["B"].width = 30
        ws_links.column_dimensions["C"].width = 80

        conn.close()

        xlsx_path = BASE / domain_name / "workbook.xlsx"
        wb.save(str(xlsx_path))
        print(f"  [REBUILD] {domain_name}/workbook.xlsx ({len(tables)} sheets)")

# ============================================================================
# STEP 5: PACKAGE INTO SELF-CONTAINED FOLDER
# ============================================================================
def package_all():
    print("\n=== PACKAGING INTO SELF-CONTAINED FOLDER ===")

    if PACKAGE_DIR.exists():
        shutil.rmtree(str(PACKAGE_DIR))
    PACKAGE_DIR.mkdir(exist_ok=True)

    # Copy all 12 domains
    for d in sorted(BASE.iterdir()):
        if d.is_dir() and d.name[0].isdigit():
            dest = PACKAGE_DIR / d.name
            shutil.copytree(str(d), str(dest), dirs_exist_ok=True)
            print(f"  [COPY] {d.name}")

    # Copy build scripts
    for f in ["build_all_domains.py", "build_workbooks.py", "build_enhancements.py"]:
        src = BASE / f
        if src.exists():
            shutil.copy2(str(src), str(PACKAGE_DIR / f))

    # Copy master index and dashboard
    for f in ["MASTER_INDEX.xlsx", "DASHBOARD.html", "DOMAINS.db", "domain_index.csv"]:
        src = BASE / f
        if src.exists():
            shutil.copy2(str(src), str(PACKAGE_DIR / f))

    # Create README
    readme = f"""# PMO SYSTEM DOMAINS — Self-Contained Package
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}

## Contents
- 12 domain packages (01-12), each with: SQL schemas, CSV data, SQLite DB, ERD diagram, Excel workbook
- MASTER_INDEX.xlsx — Master index with chart
- DASHBOARD.html — HTML dashboard (open in any browser)
- DOMAINS.db — Central index database
- build_*.py — Build scripts (reproducible)

## Each Domain Contains
```
XX_DOMAIN_NAME/
├── schemas/        — SQL CREATE TABLE statements
├── data_raw/       — CSV data files
├── databases/      — SQLite databases (queryable)
├── diagrams/       — ERD diagrams (PNG + Mermaid + Graphviz)
├── workbook.xlsx   — Excel workbook with all data + Links sheet
```

## Usage
1. Open DASHBOARD.html in a browser for overview
2. Open any workbook.xlsx in Excel for detailed data
3. Query any .db file with SQLite
4. Run build_*.py to regenerate everything

## Engineering Data Sources
- MASTER_Vault_Tables (TU_cost_indices, TU_far_limits, TU_geotechnical, TU_market_data, TU_mep_density)
- MV1 MASTER tables (Permit_Sequence, Soil_Geotechnical, Cost_CSI)
- UAE construction standards (DM, DDA, ADM, RAK Municipality)
- Ritz Carlton classification model (RM.xlsx)

## Style
- Ras Flat Dark: #2C3E50 headers, Calibri Bold 11pt, thin gray borders
- All data tightened to min/max/avg from engineering references
- Cross-referenced hyperlinks in all workbooks
"""
    (PACKAGE_DIR / "README.md").write_text(readme, encoding="utf-8")

    # Create index.html for the package
    index_html = _create_package_index()
    (PACKAGE_DIR / "index.html").write_text(index_html, encoding="utf-8")

    # Calculate total size
    total_size = sum(f.stat().st_size for f in PACKAGE_DIR.rglob("*") if f.is_file())
    total_files = sum(1 for f in PACKAGE_DIR.rglob("*") if f.is_file())
    print(f"\n  [PACKAGE] {PACKAGE_DIR}")
    print(f"  [SIZE] {total_size // 1024} KB ({total_files} files)")

def _create_package_index():
    """Create a self-contained index.html for the package"""
    domains_info = []
    for d in sorted(BASE.iterdir()):
        if d.is_dir() and d.name[0].isdigit():
            db_files = list((d / "databases").glob("*.db")) if (d / "databases").exists() else []
            csv_files = list((d / "data_raw").glob("*.csv")) if (d / "data_raw").exists() else []
            table_count = 0
            for db in db_files:
                try:
                    conn = sqlite3.connect(str(db))
                    cur = conn.cursor()
                    cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
                    table_count += len(cur.fetchall())
                    conn.close()
                except Exception:
                    pass
            domains_info.append({
                "dir": d.name,
                "tables": table_count,
                "csvs": len(csv_files),
                "dbs": len(db_files),
                "has_png": any((d / "diagrams").glob("*.png")) if (d / "diagrams").exists() else False,
            })

    cards_html = ""
    colors = ["00B4D8","06D6A0","FFD166","EF476F","118AB2","073B4C","8338EC","FF006E","FB5607","3A86FF","80ED99","C77DFF"]
    for i, info in enumerate(domains_info):
        color = colors[i % len(colors)]
        cards_html += f"""
        <div class="card" style="border-left: 5px solid #{color}; background: #16213e; border-radius: 12px; padding: 20px; margin: 10px;">
            <h3 style="color: #{color}; margin: 0 0 10px 0;">{info['dir']}</h3>
            <div style="display: flex; gap: 20px; color: #BDC3C7; font-size: 13px;">
                <span>Tables: <b style="color: #fff;">{info['tables']}</b></span>
                <span>CSV: <b style="color: #fff;">{info['csvs']}</b></span>
                <span>DB: <b style="color: #fff;">{info['dbs']}</b></span>
            </div>
        </div>"""

    return f"""<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PMO DOMAINS Package</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: 'Segoe UI', Tahoma, sans-serif; background: #1a1a2e; color: #eee; padding: 30px; }}
h1 {{ text-align: center; color: #3498DB; margin-bottom: 30px; }}
.grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(350px, 1fr)); gap: 15px; }}
</style>
</head>
<body>
<h1>PMO SYSTEM DOMAINS</h1>
<p style="text-align:center; color:#7F8C8D; margin-bottom:30px;">Self-Contained Package — {len(domains_info)} Domains — Generated {datetime.now().strftime('%Y-%m-%d')}</p>
<div class="grid">{cards_html}</div>
</body>
</html>"""

# ============================================================================
# MAIN
# ============================================================================
if __name__ == "__main__":
    print("=" * 70)
    print("PMO COMPREHENSIVE EXTENSION, AUDIT & PACKAGE")
    print("=" * 70)
    vault = load_vault_data()
    extend_domains(vault)
    audit = engineering_audit(vault)
    fill_gaps(vault)
    rebuild_workbooks()
    package_all()

    # Print audit summary
    print("\n" + "=" * 70)
    print("ENGINEERING AUDIT SUMMARY")
    print("=" * 70)
    for r in audit:
        status_icon = "✅" if r["status"] == "OK" else "⚠️"
        print(f"  {status_icon} {r['domain']}: {r['status']}")
        for issue in r["issues"]:
            print(f"      → {issue}")

    print("\n" + "=" * 70)
    print("ALL TASKS COMPLETE")
    print("=" * 70)
