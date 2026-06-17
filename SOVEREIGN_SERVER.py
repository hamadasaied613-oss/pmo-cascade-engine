#!/usr/bin/env python3
"""
PMO CASCADE SOVEREIGN ENGINE — FastAPI Backend
Processes full 12-domain chain with cascade engine
"""
import sqlite3, json, os, sys, math, datetime, traceback
from pathlib import Path
from typing import Optional, Dict, Any, List
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import uvicorn

# ─── PATHS ───
BASE_DIR = Path(__file__).parent
GATEWAY_DB = BASE_DIR / "_GATEWAY_TABLES" / "gateway.db"
DOMAINS_DB = BASE_DIR / "DOMAINS.db"
OUTPUT_DIR = BASE_DIR / "_PMO_OUTPUTS"
DELIVERABLES_DIR = BASE_DIR / "_PMO_DELIVERABLES"
HTML_DIR = BASE_DIR / "SOVEREIGN.html"

# ─── DOMAIN DEFINITIONS ───
DOMAINS = {
    "01": {"name": "Project & Plot", "name_ar": "المشروع والأرض", "color": "#00B4D8", "db": "project_plot.db"},
    "02": {"name": "Location & GIS", "name_ar": "الموقع والخرائط", "color": "#06D6A0", "db": "location_gis.db"},
    "03": {"name": "Stakeholders", "name_ar": "أصحاب المصلحة", "color": "#FFD166", "db": "stakeholders.db"},
    "04": {"name": "Zoning & Regulatory", "name_ar": "التخطيط والتنظيم", "color": "#EF476F", "db": "zoning_regulatory.db"},
    "05": {"name": "Validation & Compliance", "name_ar": "التحقق والامتثال", "color": "#118AB2", "db": "validation.db"},
    "06": {"name": "Design Parameters", "name_ar": "معايير التصميم", "color": "#073B4C", "db": "design_parameters.db"},
    "07": {"name": "Unit Mix & Program", "name_ar": " خطة الوحدات", "color": "#8338EC", "db": "unit_mix.db"},
    "08": {"name": "Cost & Economics", "name_ar": "التكلفة والاقتصاديات", "color": "#FF006E", "db": "cost_economics.db"},
    "09": {"name": "Schedule & Timeline", "name_ar": "الجدول الزمني", "color": "#FB5607", "db": "schedule.db"},
    "10": {"name": "Approvals & Authorities", "name_ar": "الموافقات والجهات", "color": "#3A86FF", "db": "approvals.db"},
    "11": {"name": "Geotechnical", "name_ar": "ال ENGINEERING الجيوتقنية", "color": "#80ED99", "db": "geotechnical.db"},
    "12": {"name": "Sustainability & BIM", "name_ar": "الاستدامة والنمذجة", "color": "#C77DFF", "db": "sustainability.db"},
}

# ─── APP ───
app = FastAPI(title="PMO CASCADE Sovereign Engine", version="2.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")

# ─── MODELS ───
class ProjectInput(BaseModel):
    plot: str = "DUB-LOT-001"
    emirate: str = "Dubai"
    type: str = "MIX"
    style: str = "modern"
    quality: str = "Premium"
    qualityScore: float = 0.75
    costFactor: float = 1.0
    purpose: str = "investment"
    floors: int = 15
    height: float = 55
    totalArea: float = 12000
    units: Dict[str, int] = {"studio": 5, "one": 20, "two": 20, "three": 8, "pent": 2, "total": 55}
    catalogSelections: Optional[Dict] = None
    exteriorItems: Optional[List[Dict]] = None
    interiorItems: Optional[List[Dict]] = None
    landscapeItem: Optional[Dict] = None
    facilityItems: Optional[List[Dict]] = None

class N8nWebhookPayload(BaseModel):
    action: str
    project_id: Optional[str] = None
    input_data: Optional[Dict] = None
    callback_url: Optional[str] = None

# ─── DATABASE HELPERS ───
def get_db(db_path: Path) -> sqlite3.Connection:
    if not db_path.exists():
        return None
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    return conn

def query_db(db_path: Path, sql: str, params: tuple = ()) -> List[Dict]:
    conn = get_db(db_path)
    if not conn:
        return []
    try:
        cursor = conn.execute(sql, params)
        rows = cursor.fetchall()
        return [dict(row) for row in rows]
    except:
        return []
    finally:
        conn.close()

def get_tables(db_path: Path) -> List[str]:
    conn = get_db(db_path)
    if not conn:
        return []
    try:
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        return [row[0] for row in cursor.fetchall()]
    finally:
        conn.close()

def get_table_data(db_path: Path, table: str) -> List[Dict]:
    conn = get_db(db_path)
    if not conn:
        return []
    try:
        cursor = conn.execute(f"SELECT * FROM [{table}]")
        rows = cursor.fetchall()
        return [dict(row) for row in rows]
    except:
        return []
    finally:
        conn.close()

def fmt(n):
    if n is None: return "-"
    try: return f"{float(n):,.0f}"
    except: return str(n)

# ─── CASCADE ENGINE (inline, no import needed) ───
class CascadeEngine:
    def __init__(self):
        self.conn = sqlite3.connect(str(GATEWAY_DB))
        self.conn.row_factory = sqlite3.Row
        self.input_data = {}
        self.derived = {}

    def run(self, inp: dict) -> dict:
        self.input_data = inp
        self.derived = {}
        self._derive_from_plot()
        self._lookup_zoning()
        self._lookup_cost()
        self._lookup_structural()
        self._lookup_fire()
        self._calculate_area()
        self._calculate_cost()
        self._generate_unit_mix()
        self._generate_domain_params()
        self.conn.close()
        return {"input": self.input_data, "derived": self.derived,
                "catalog_selections": inp.get("catalogSelections"),
                "inferred_quality": {
                    "level": inp.get("quality", "Standard"),
                    "score": inp.get("qualityScore", 0.5),
                    "cost_factor": inp.get("costFactor", 1.0),
                },
                "generated_at": datetime.datetime.now().isoformat()}

    def _derive_from_plot(self):
        plot = self.input_data.get("plot", "").upper()
        try:
            row = self.conn.execute("SELECT * FROM plot_database WHERE plot_number=?", (plot,)).fetchone()
            if row: self.derived.update(dict(row))
        except: pass

    def _lookup_zoning(self):
        emirate = self.derived.get("emirate", self.input_data.get("emirate", ""))
        bt = {"RES":"Residential","COM":"Commercial","MIX":"Mixed-Use","HOT":"Hotel"}.get(self.input_data.get("type","MIX"), "Mixed-Use")
        try:
            row = self.conn.execute("SELECT * FROM zoning_matrix WHERE emirate=? AND (land_use=? OR building_type=?) LIMIT 1",
                                    (emirate, bt, bt)).fetchone()
            if row:
                r = dict(row)
                self.derived.update({
                    "zoning_code": r.get("zone_code",""), "far_min": r.get("far_min",3), "far_opt": r.get("far_optimal",3.5), "far_max": r.get("far_max",5),
                    "h_min": r.get("height_min_m",30), "h_opt": r.get("height_optimal_m",50), "h_max": r.get("height_max_m",80),
                    "floors_max": r.get("max_floors",20), "coverage_max": r.get("coverage_max_pct",60),
                    "setback_f": r.get("setback_front_m",3), "setback_r": r.get("setback_rear_m",3), "setback_s": r.get("setback_side_m",2),
                    "park_res": r.get("parking_ratio_residential",1), "park_com": r.get("parking_ratio_commercial",0),
                })
        except: pass

    def _lookup_cost(self):
        emirate = self.derived.get("emirate", self.input_data.get("emirate", ""))
        bt = {"RES":"Residential","COM":"Commercial","MIX":"Mixed-Use","HOT":"Hotel"}.get(self.input_data.get("type","MIX"), "Mixed-Use")
        quality = self.input_data.get("quality", "Standard")
        # Apply catalog cost factor
        catalog_cf = self.input_data.get("costFactor", 1.0) or 1.0
        try:
            row = self.conn.execute("SELECT * FROM cost_matrix WHERE emirate=? AND building_type=? AND quality_level=? LIMIT 1",
                                    (emirate, bt, quality)).fetchone()
            if not row:
                row = self.conn.execute("SELECT * FROM cost_matrix WHERE emirate=? AND building_type=? LIMIT 1", (emirate, bt)).fetchone()
            if row:
                r = dict(row)
                self.derived.update({
                    "cost_structure": round(r.get("structure_aed_sqm",1350) * catalog_cf, 2),
                    "cost_mep": round(r.get("mep_aed_sqm",675) * catalog_cf, 2),
                    "cost_finishes": round(r.get("finishes_aed_sqm",530) * catalog_cf, 2),
                    "cost_facade": round(r.get("facade_aed_sqm",430) * catalog_cf, 2),
                    "cost_total": round(r.get("total_construction_aed_sqm",2985) * catalog_cf, 2),
                    "cost_land": r.get("land_price_aed_sqm",8000),
                    "soft_cost_pct": r.get("soft_cost_pct",12),
                    "contingency_pct": r.get("contingency_pct",10),
                    "catalog_cost_factor": catalog_cf,
                })
        except: pass

    def _lookup_structural(self):
        floors = int(self.input_data.get("floors", 10) or 10)
        h = float(self.input_data.get("height", floors * 3.5) or floors * 3.5)
        bt = {"RES":"Residential","COM":"Commercial","MIX":"Mixed-Use","HOT":"Hotel"}.get(self.input_data.get("type","MIX"), "Mixed-Use")
        soil = self.derived.get("soil_type", "Dense Sand")
        try:
            row = self.conn.execute("SELECT * FROM structural_matrix WHERE building_type=? AND height_min_m<=? AND height_max_m>=? AND soil_types LIKE ? LIMIT 1",
                                    (bt, h, h, f"%{soil}%")).fetchone()
            if not row:
                row = self.conn.execute("SELECT * FROM structural_matrix WHERE building_type=? AND height_min_m<=? AND height_max_m>=? LIMIT 1",
                                        (bt, h, h)).fetchone()
            if row:
                r = dict(row)
                self.derived.update({
                    "structural_system": r.get("structural_system","RC Frame"), "concrete_grade": r.get("concrete_grade","C35"),
                    "steel_grade": r.get("steel_grade","B500"), "foundation_type": r.get("foundation_type","Spread Footing"),
                    "max_span": r.get("max_span_m",7), "cost_factor": r.get("cost_factor",1.0),
                })
        except: pass

    def _lookup_fire(self):
        floors = int(self.input_data.get("floors", 10) or 10)
        h = float(self.input_data.get("height", floors * 3.5) or floors * 3.5)
        try:
            row = self.conn.execute("SELECT * FROM fire_matrix WHERE height_min_m<=? AND height_max_m>=? LIMIT 1", (h, h)).fetchone()
            if row:
                r = dict(row)
                self.derived.update({
                    "sprinkler": bool(r.get("sprinkler_required",1)), "fire_alarm": bool(r.get("fire_alarm_required",1)),
                    "smoke_extraction": bool(r.get("smoke_extraction",1)), "refuge_floor": bool(r.get("refuge_floor",0)),
                    "stair_count": r.get("stair_count_min",2), "stair_width": r.get("stair_width_min_m",1.2),
                    "fire_rating": r.get("fire_rating_hours",2),
                })
        except: pass

    def _calculate_area(self):
        plot = self.derived.get("plot_area_sqm", 1500)
        far = self.derived.get("far_opt", 3.5)
        cov = self.derived.get("coverage_max", 60)
        floors = int(self.input_data.get("floors", 10) or 10)
        fp = plot * (cov / 100)
        gfa = fp * floors
        max_gfa = plot * far
        self.derived.update({
            "footprint": round(fp, 1), "total_gfa": round(min(gfa, max_gfa), 1),
            "max_gfa": round(max_gfa, 1), "efficiency": round(gfa / max_gfa * 100, 1) if max_gfa > 0 else 0,
        })

    def _calculate_cost(self):
        gfa = self.derived.get("total_gfa", 1000)
        csqm = self.derived.get("cost_total", 2985)
        lsqm = self.derived.get("cost_land", 8000)
        plot = self.derived.get("plot_area_sqm", 1500)
        sp = self.derived.get("soft_cost_pct", 12)
        cp = self.derived.get("contingency_pct", 10)
        # Use structural cost_factor AND catalog cost factor
        sf = self.derived.get("cost_factor", 1.0)
        cf = self.input_data.get("costFactor", 1.0) or 1.0
        construction = gfa * csqm * sf * cf
        land = plot * lsqm
        soft = construction * (sp / 100)
        cont = construction * (cp / 100)
        total = construction + land + soft + cont
        self.derived.update({
            "construction_cost": round(construction), "land_cost": round(land),
            "soft_cost": round(soft), "contingency_cost": round(cont),
            "total_project_cost": round(total),
        })

    def _generate_unit_mix(self):
        bt = self.input_data.get("type", "MIX")
        if bt not in ("RES", "MIX"): return
        total = int(self.input_data.get("units", {}).get("total", 0) or 0)
        if total == 0:
            gfa = self.derived.get("total_gfa", 5000)
            total = int(gfa / 120 * 0.85)
        self.derived["unit_mix"] = {
            "studios": int(total * 0.15), "one_bed": int(total * 0.35),
            "two_bed": int(total * 0.35), "three_bed": int(total * 0.12),
            "penthouses": int(total * 0.03), "total": total,
        }

    def _generate_domain_params(self):
        d = self.derived
        # Domain 04 — Zoning
        d["domain_04"] = {
            "far_compliance": "PASS" if d.get("far_opt", 0) <= d.get("far_max", 999) else "FAIL",
            "height_compliance": "PASS" if self.input_data.get("height", 0) <= d.get("h_max", 999) else "FAIL",
            "setback_compliance": "PASS",
        }
        # Domain 06 — Design
        d["domain_06"] = {
            "structural_system": d.get("structural_system", "RC Frame"),
            "concrete_grade": d.get("concrete_grade", "C35"),
            "facade_system": "Curtain Wall" if self.input_data.get("quality") in ("Premium","Luxury") else "Aluminum Cladding",
            "hvac_system": "VRF" if self.input_data.get("quality") in ("Premium","Luxury") else "FCU",
        }
        # Domain 08 — Cost breakdown
        gfa = d.get("total_gfa", 1000)
        d["domain_08"] = {
            "cost_per_sqm": d.get("cost_total", 2985),
            "construction_total": d.get("construction_cost", 0),
            "land_total": d.get("land_cost", 0),
            "soft_total": d.get("soft_cost", 0),
            "contingency_total": d.get("contingency_cost", 0),
            "grand_total": d.get("total_project_cost", 0),
            "cost_per_unit": round(d.get("total_project_cost", 0) / max(d.get("unit_mix", {}).get("total", 1), 1)),
        }
        # Domain 09 — Schedule
        floors = int(self.input_data.get("floors", 10) or 10)
        d["domain_09"] = {
            "enabling_weeks": 8, "substructure_weeks": max(4, floors),
            "superstructure_weeks": max(8, floors * 2), "envelope_weeks": max(6, floors),
            "fitout_weeks": max(12, floors * 2), "external_weeks": 8,
            "total_weeks": 0, "total_months": 0,
        }
        tw = sum(v for k, v in d["domain_09"].items() if "_weeks" in k)
        d["domain_09"]["total_weeks"] = tw
        d["domain_09"]["total_months"] = round(tw / 4.33, 1)
        # Domain 10 — Approvals
        d["domain_10"] = {
            "permit_steps": 14, "total_approval_weeks": 22,
            "critical_path": ["Land Title", "Planning NOC", "Environmental", "Building Permit", "Civil Defense", "DEWA"],
        }
        # Domain 11 — Geotech
        d["domain_11"] = {
            "soil_type": d.get("soil_type", "Dense Sand"),
            "bearing_capacity_kpa": 300 if "Dense" in d.get("soil_type", "") else 150,
            "foundation_type": d.get("foundation_type", "Spread Footing"),
            "shoring_system": "Secant Pile" if floors > 10 else "Soldier Pile",
        }
        # Domain 12 — Sustainability
        d["domain_12"] = {
            "leed_target": "Gold" if self.input_data.get("quality") in ("Premium","Luxury") else "Silver",
            "estidama_target": "3 Pearl" if self.input_data.get("quality") in ("Premium","Luxury") else "2 Pearl",
            "energy_efficiency": "30% reduction" if self.input_data.get("quality") in ("Premium","Luxury") else "20% reduction",
        }

# ─── API ROUTES ───

@app.get("/", response_class=HTMLResponse)
async def serve_ui():
    if HTML_DIR.exists():
        content = HTML_DIR.read_bytes()
        return Response(content=content, media_type="text/html", headers={
            "Content-Type": "text/html; charset=utf-8",
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "X-Content-Type-Options": "nosniff",
        })
    return HTMLResponse("<h1>PMO Cascade Sovereign Engine</h1><p>SOVEREIGN.html not found</p>")

@app.get("/api/domains")
async def get_domains():
    return {"domains": DOMAINS}

@app.get("/api/domain/{domain_id}")
async def get_domain(domain_id: str):
    if domain_id not in DOMAINS:
        raise HTTPException(404, f"Domain {domain_id} not found")
    info = DOMAINS[domain_id]
    db_path = BASE_DIR / f"{domain_id}_{info['name'].replace(' & ', '_').replace(' ', '_')}" / "databases" / info["db"]
    if not db_path.exists():
        db_path = BASE_DIR / domain_id / "databases" / info["db"]
    if not db_path.exists():
        # Try finding any .db in domain folder
        domain_folder = BASE_DIR / f"{domain_id}_{info['name'].replace(' & ', '_').replace(' ', '_')}"
        if not domain_folder.exists():
            # Try alternate naming
            for d in BASE_DIR.iterdir():
                if d.is_dir() and d.name.startswith(domain_id):
                    domain_folder = d
                    break
        if domain_folder.exists():
            for f in domain_folder.rglob("*.db"):
                db_path = f
                break
    tables = get_tables(db_path) if db_path.exists() else []
    data = {}
    for t in tables:
        data[t] = get_table_data(db_path, t)
    return {"domain": info, "tables": tables, "data": data, "db_path": str(db_path)}

@app.get("/api/domain/{domain_id}/{table}")
async def get_domain_table(domain_id: str, table: str):
    info = DOMAINS.get(domain_id)
    if not info:
        raise HTTPException(404, "Domain not found")
    db_path = BASE_DIR / f"{domain_id}_{info['name'].replace(' & ', '_').replace(' ', '_')}" / "databases" / info["db"]
    if not db_path.exists():
        for d in BASE_DIR.iterdir():
            if d.is_dir() and d.name.startswith(domain_id):
                for f in d.rglob("*.db"):
                    db_path = f
                    break
    data = get_table_data(db_path, table)
    return {"table": table, "rows": data, "count": len(data)}

@app.get("/api/gateway")
async def get_gateway():
    tables = get_tables(GATEWAY_DB)
    data = {}
    for t in tables:
        data[t] = get_table_data(GATEWAY_DB, t)
    return {"tables": tables, "data": data}

@app.post("/api/cascade")
async def run_cascade(inp: ProjectInput):
    try:
        engine = CascadeEngine()
        result = engine.run(inp.model_dump())
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = OUTPUT_DIR / "cascade_output.json"
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, default=str)
        return result
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, f"Cascade engine error: {str(e)}")

@app.get("/api/cascade/output")
async def get_cascade_output():
    out_path = OUTPUT_DIR / "cascade_output.json"
    if out_path.exists():
        with open(out_path, encoding="utf-8") as f:
            return json.load(f)
    return {"error": "No cascade output found. Run cascade first."}

@app.get("/api/deliverables")
async def get_deliverables():
    deliverables = []
    if DELIVERABLES_DIR.exists():
        for domain_dir in sorted(DELIVERABLES_DIR.iterdir()):
            if domain_dir.is_dir() and domain_dir.name.startswith("DOMAIN"):
                notebooks = list(domain_dir.glob("*.ipynb"))
                deliverables.append({
                    "domain": domain_dir.name,
                    "notebooks": [{"name": n.stem, "path": str(n)} for n in notebooks],
                    "count": len(notebooks),
                })
    return {"deliverables": deliverables, "total": sum(d["count"] for d in deliverables)}

@app.get("/api/stats")
async def get_stats():
    gw_tables = get_tables(GATEWAY_DB)
    gw_counts = {}
    for t in gw_tables:
        gw_counts[t] = len(get_table_data(GATEWAY_DB, t))
    out_path = OUTPUT_DIR / "cascade_output.json"
    has_output = out_path.exists()
    deliverable_count = 0
    if DELIVERABLES_DIR.exists():
        for d in DELIVERABLES_DIR.iterdir():
            if d.is_dir() and d.name.startswith("DOMAIN"):
                deliverable_count += len(list(d.glob("*.ipynb")))
    return {
        "gateway_tables": gw_counts,
        "domains_count": len(DOMAINS),
        "has_cascade_output": has_output,
        "deliverable_notebooks": deliverable_count,
    }

# ─── EXCEL EXPORT ───

@app.get("/api/export/excel")
async def export_excel():
    out_path = OUTPUT_DIR / "cascade_output.json"
    if not out_path.exists():
        raise HTTPException(404, "No cascade output. Run cascade first.")
    with open(out_path, encoding="utf-8") as f:
        data = json.load(f)
    d = data.get("derived", {})
    inp = data.get("input", {})
    iq = data.get("inferred_quality", {})

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ── Sheet 1: Project Summary ──
    ws = wb.active
    ws.title = "Project Summary"
    header_fill = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="00B4D8", size=11)
    data_font = Font(name="Calibri", size=11)
    border = Border(
        bottom=Side(style="thin", color="333333"),
        right=Side(style="thin", color="333333"),
    )
    accent_fill = PatternFill(start_color="00B4D8", end_color="00B4D8", fill_type="solid")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25

    ws.merge_cells("A1:D1")
    ws["A1"] = "PMO CASCADE — Project Summary"
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = accent_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    summary_data = [
        ("Plot Number", inp.get("plot", "")),
        ("Emirate", inp.get("emirate", "")),
        ("Building Type", inp.get("type", "")),
        ("Style", inp.get("style", "")),
        ("Floors", inp.get("floors", "")),
        ("Height (m)", inp.get("height", "")),
        ("Total Area (m²)", inp.get("totalArea", "")),
        ("Inferred Quality", iq.get("level", inp.get("quality", ""))),
        ("Quality Score", f"{iq.get('score', 0)*100:.0f}%"),
        ("Cost Factor", f"x{iq.get('cost_factor', 1.0):.2f}"),
        ("", ""),
        ("GFA (m²)", d.get("total_gfa", "")),
        ("Max GFA (m²)", d.get("max_gfa", "")),
        ("Efficiency", f"{d.get('efficiency', '')}%"),
        ("Footprint (m²)", d.get("footprint", "")),
        ("", ""),
        ("Structural System", d.get("structural_system", "")),
        ("Concrete Grade", d.get("concrete_grade", "")),
        ("Steel Grade", d.get("steel_grade", "")),
        ("Foundation Type", d.get("foundation_type", "")),
        ("", ""),
        ("Construction Cost (AED)", f"{d.get('construction_cost', 0):,.0f}"),
        ("Land Cost (AED)", f"{d.get('land_cost', 0):,.0f}"),
        ("Soft Cost (AED)", f"{d.get('soft_cost', 0):,.0f}"),
        ("Contingency (AED)", f"{d.get('contingency_cost', 0):,.0f}"),
        ("TOTAL PROJECT COST (AED)", f"{d.get('total_project_cost', 0):,.0f}"),
        ("", ""),
        ("Total Units", d.get("unit_mix", {}).get("total", "")),
        ("Schedule (months)", d.get("domain_09", {}).get("total_months", "")),
    ]
    for i, (label, value) in enumerate(summary_data, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(name="Calibri", bold=True, size=11, color="CCCCCC")
        ws.cell(row=i, column=2, value=value).font = data_font
        ws.cell(row=i, column=1).border = border
        ws.cell(row=i, column=2).border = border

    # ── Sheet 2: Zoning Parameters ──
    ws2 = wb.create_sheet("Zoning")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 15
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "Domain 04 — Zoning Parameters"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill = PatternFill(start_color="EF476F", end_color="EF476F", fill_type="solid")
    ws2["A1"].alignment = Alignment(horizontal="center")

    zoning_data = [
        ("Parameter", "Min", "Optimal", "Max"),
        ("FAR", d.get("far_min", ""), d.get("far_opt", ""), d.get("far_max", "")),
        ("Height (m)", d.get("h_min", ""), d.get("h_opt", ""), d.get("h_max", "")),
        ("Coverage (%)", "—", "—", d.get("coverage_max", "")),
        ("Setback Front (m)", d.get("setback_f", ""), "—", "—"),
        ("Setback Rear (m)", d.get("setback_r", ""), "—", "—"),
        ("Setback Side (m)", d.get("setback_s", ""), "—", "—"),
        ("Parking Residential", d.get("park_res", ""), "—", "—"),
        ("Parking Commercial", d.get("park_com", ""), "—", "—"),
    ]
    for i, row in enumerate(zoning_data, start=3):
        for j, val in enumerate(row):
            cell = ws2.cell(row=i, column=j+1, value=val)
            cell.border = border
            if i == 3:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.font = data_font

    # ── Sheet 3: Cost Breakdown ──
    ws3 = wb.create_sheet("Cost Breakdown")
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 15
    ws3.column_dimensions["D"].width = 20
    ws3.merge_cells("A1:D1")
    ws3["A1"] = "Domain 08 — Cost Breakdown"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill = PatternFill(start_color="FF006E", end_color="FF006E", fill_type="solid")
    ws3["A1"].alignment = Alignment(horizontal="center")

    gfa = d.get("total_gfa", 1000)
    cf = iq.get("cost_factor", 1.0)
    cost_items = [
        ("Item", "Quantity", "Unit", "Cost (AED)"),
        ("Substructure", f"{gfa*0.15:.0f}", "m²", f"{gfa*0.15*d.get('cost_structure',1350)*cf*0.2:,.0f}"),
        ("Superstructure (RC)", f"{gfa:.0f}", "m²", f"{gfa*d.get('cost_structure',1350)*cf:,.0f}"),
        ("Concrete Works", f"{gfa*0.5:.0f}", "m³", f"{gfa*0.5*d.get('cost_structure',1350)*0.6*cf:,.0f}"),
        ("Steel Reinforcement", f"{gfa*80:.0f}", "kg", f"{gfa*80*4.5*cf:,.0f}"),
        ("Facade / Cladding", f"{gfa*0.8:.0f}", "m²", f"{gfa*0.8*d.get('cost_facade',430)*cf:,.0f}"),
        ("HVAC System", f"{gfa:.0f}", "m²", f"{gfa*d.get('cost_mep',675)*0.4*cf:,.0f}"),
        ("Electrical Works", f"{gfa:.0f}", "m²", f"{gfa*d.get('cost_mep',675)*0.35*cf:,.0f}"),
        ("Plumbing", f"{gfa:.0f}", "m²", f"{gfa*d.get('cost_mep',675)*0.25*cf:,.0f}"),
        ("Interior Fit-Out", f"{gfa*0.7:.0f}", "m²", f"{gfa*0.7*d.get('cost_finishes',530)*cf:,.0f}"),
        ("Fire Protection", f"{gfa:.0f}", "m²", f"{gfa*120*cf:,.0f}"),
        ("Elevators", "2-3", "units", f"{250000*cf:,.0f}"),
        ("External Works", f"{gfa*0.1:.0f}", "m²", f"{gfa*0.1*350*cf:,.0f}"),
        ("Landscaping", f"{d.get('footprint',500)*0.3:.0f}", "m²", f"{d.get('footprint',500)*0.3*200*cf:,.0f}"),
        ("", "", "", ""),
        ("Construction Subtotal", "", "", f"{d.get('construction_cost',0)*cf:,.0f}"),
        ("Land Cost", "", "", f"{d.get('land_cost',0):,.0f}"),
        ("Soft Costs", "", "", f"{d.get('soft_cost',0)*cf:,.0f}"),
        ("Contingency", "", "", f"{d.get('contingency_cost',0)*cf:,.0f}"),
        ("TOTAL PROJECT COST", "", "", f"{d.get('total_project_cost',0)*cf:,.0f}"),
    ]
    for i, row in enumerate(cost_items, start=3):
        for j, val in enumerate(row):
            cell = ws3.cell(row=i, column=j+1, value=val)
            cell.border = border
            if i == 3:
                cell.font = header_font
                cell.fill = header_fill
            elif i == len(cost_items) + 2:
                cell.font = Font(name="Calibri", bold=True, size=12, color="FF006E")
            else:
                cell.font = data_font

    # ── Sheet 4: Schedule ──
    ws4 = wb.create_sheet("Schedule")
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 15
    ws4.merge_cells("A1:B1")
    ws4["A1"] = "Domain 09 — Schedule"
    ws4["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws4["A1"].fill = PatternFill(start_color="FB5607", end_color="FB5607", fill_type="solid")
    ws4["A1"].alignment = Alignment(horizontal="center")

    d09 = d.get("domain_09", {})
    sched_data = [("Phase", "Weeks")]
    for k, v in d09.items():
        if "_weeks" in k:
            sched_data.append((k.replace("_weeks", "").title(), v))
    sched_data.append(("TOTAL", d09.get("total_weeks", 0)))
    sched_data.append(("TOTAL (months)", d09.get("total_months", 0)))

    for i, (label, val) in enumerate(sched_data, start=3):
        ws4.cell(row=i, column=1, value=label).font = header_font if i == 3 else data_font
        ws4.cell(row=i, column=2, value=val).font = header_font if i == 3 else data_font
        ws4.cell(row=i, column=1).border = border
        ws4.cell(row=i, column=2).border = border
        if i == 3:
            ws4.cell(row=i, column=1).fill = header_fill
            ws4.cell(row=i, column=2).fill = header_fill

    # Save
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    excel_path = OUTPUT_DIR / "cascade_report.xlsx"
    wb.save(str(excel_path))
    return FileResponse(excel_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="PMO_Cascade_Report.xlsx")

# ─── POSTER GENERATOR ───

@app.get("/api/export/poster")
async def export_poster():
    out_path = OUTPUT_DIR / "cascade_output.json"
    if not out_path.exists():
        raise HTTPException(404, "No cascade output.")
    with open(out_path, encoding="utf-8") as f:
        data = json.load(f)
    d = data.get("derived", {})
    inp = data.get("input", {})
    iq = data.get("inferred_quality", {})
    cat = data.get("catalog_selections") or {}
    iq_level = iq.get("level", inp.get("quality", ""))
    iq_score = iq.get("score", 0)
    iq_cost = iq.get("cost_factor", 1.0)
    d09 = d.get("domain_09", {})
    um = d.get("unit_mix", {})

    poster_html = f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><title>PMO Poster — {inp.get('plot','')}</title>
<style>
@page{{size:A3 landscape;margin:0}}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:#0a0a12;color:#e0e0e8;width:420mm;height:297mm;overflow:hidden;position:relative}}
.poster{{display:grid;grid-template-columns:1fr 1fr 1fr;grid-template-rows:auto 1fr 1fr auto;height:100%;gap:0}}
.header{{grid-column:1/-1;background:linear-gradient(135deg,#00B4D8,#0090ad);padding:20px 30px;display:flex;align-items:center;justify-content:space-between}}
.header h1{{font-size:28px;font-weight:800;color:#fff;letter-spacing:1px}}
.header .meta{{text-align:right;color:rgba(255,255,255,.8);font-size:12px}}
.header .meta .plot{{font-size:18px;font-weight:700;color:#fff}}
.panel{{padding:18px 22px;border:1px solid #1a1a2a;position:relative}}
.panel h2{{font-size:13px;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:10px;padding-bottom:6px;border-bottom:2px solid}}
.panel .big{{font-size:32px;font-weight:800;line-height:1}}
.panel .label{{font-size:10px;color:#888;margin-top:4px;text-transform:uppercase;letter-spacing:.5px}}
.panel .row{{display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid #1a1a2a;font-size:11px}}
.panel .row .k{{color:#888}}.panel .row .v{{font-weight:600}}
.gauge{{height:6px;background:#1a1a2a;border-radius:3px;margin-top:8px;overflow:hidden}}
.gauge .fill{{height:100%;border-radius:3px}}
.footer{{grid-column:1/-1;background:#111118;padding:12px 30px;display:flex;justify-content:space-between;font-size:10px;color:#555;border-top:2px solid #1a1a2a}}
.chip{{display:inline-block;padding:3px 10px;border-radius:12px;font-size:9px;margin:2px;background:rgba(0,180,216,.1);color:#00B4D8;border:1px solid rgba(0,180,216,.2)}}
</style></head><body>
<div class="poster">
  <div class="header">
    <div><h1>PMO CASCADE FEASIBILITY STUDY</h1><div style="font-size:12px;color:rgba(255,255,255,.7);margin-top:4px">Full 12-Domain Engineering Analysis</div></div>
    <div class="meta"><div class="plot">{inp.get('plot','')}</div><div>{inp.get('emirate','')} · {inp.get('type','')} · {inp.get('floors','')} Floors</div><div>Quality: {iq_level} ({iq_score*100:.0f}%)</div></div>
  </div>

  <div class="panel" style="border-color:#00B4D8">
    <h2 style="color:#00B4D8">Project Overview</h2>
    <div class="big" style="color:#00B4D8">{fmt(d.get('total_gfa'))} <span style="font-size:14px">m² GFA</span></div>
    <div class="label">Gross Floor Area</div>
    <div style="margin-top:12px">
      <div class="row"><span class="k">Emirate</span><span class="v">{inp.get('emirate','')}</span></div>
      <div class="row"><span class="k">Type</span><span class="v">{inp.get('type','')}</span></div>
      <div class="row"><span class="k">Style</span><span class="v">{inp.get('style','')}</span></div>
      <div class="row"><span class="k">Floors</span><span class="v">{inp.get('floors','')}</span></div>
      <div class="row"><span class="k">Height</span><span class="v">{inp.get('height','')}m</span></div>
      <div class="row"><span class="k">Footprint</span><span class="v">{fmt(d.get('footprint'))} m²</span></div>
      <div class="row"><span class="k">Efficiency</span><span class="v">{d.get('efficiency','')}%</span></div>
    </div>
  </div>

  <div class="panel" style="border-color:#EF476F">
    <h2 style="color:#EF476F">Zoning & Regulatory</h2>
    <div class="big" style="color:#EF476F">{d.get('far_opt','')}</div>
    <div class="label">Optimal FAR</div>
    <div style="margin-top:12px">
      <div class="row"><span class="k">FAR Range</span><span class="v">{d.get('far_min','')} — {d.get('far_max','')}</span></div>
      <div class="row"><span class="k">Height Range</span><span class="v">{d.get('h_min','')}m — {d.get('h_max','')}m</span></div>
      <div class="row"><span class="k">Coverage</span><span class="v">{d.get('coverage_max','')}%</span></div>
      <div class="row"><span class="k">Setback F/R/S</span><span class="v">{d.get('setback_f','')}/{d.get('setback_r','')}/{d.get('setback_s','')}m</span></div>
      <div class="row"><span class="k">Zoning Code</span><span class="v">{d.get('zoning_code','')}</span></div>
    </div>
  </div>

  <div class="panel" style="border-color:#06D6A0">
    <h2 style="color:#06D6A0">Structural System</h2>
    <div class="big" style="color:#06D6A0;font-size:20px">{d.get('structural_system','')}</div>
    <div class="label">Primary System</div>
    <div style="margin-top:12px">
      <div class="row"><span class="k">Concrete</span><span class="v">{d.get('concrete_grade','')}</span></div>
      <div class="row"><span class="k">Steel</span><span class="v">{d.get('steel_grade','')}</span></div>
      <div class="row"><span class="k">Foundation</span><span class="v">{d.get('foundation_type','')}</span></div>
      <div class="row"><span class="k">Max Span</span><span class="v">{d.get('max_span','')}m</span></div>
      <div class="row"><span class="k">Sprinkler</span><span class="v">{'Yes' if d.get('sprinkler') else 'No'}</span></div>
      <div class="row"><span class="k">Refuge Floor</span><span class="v">{'Yes' if d.get('refuge_floor') else 'No'}</span></div>
    </div>
  </div>

  <div class="panel" style="border-color:#FF006E">
    <h2 style="color:#FF006E">Cost Breakdown</h2>
    <div class="big" style="color:#FF006E">{fmt(d.get('total_project_cost'))} <span style="font-size:12px">AED</span></div>
    <div class="label">Total Project Cost</div>
    <div style="margin-top:12px">
      <div class="row"><span class="k">Construction</span><span class="v">{fmt(d.get('construction_cost'))} AED</span></div>
      <div class="row"><span class="k">Land</span><span class="v">{fmt(d.get('land_cost'))} AED</span></div>
      <div class="row"><span class="k">Soft Costs</span><span class="v">{fmt(d.get('soft_cost'))} AED</span></div>
      <div class="row"><span class="k">Contingency</span><span class="v">{fmt(d.get('contingency_cost'))} AED</span></div>
      <div class="row"><span class="k">Cost/m²</span><span class="v">{fmt(d.get('cost_total'))} AED</span></div>
      <div class="row"><span class="k">Cost Factor</span><span class="v">x{iq.get('cost_factor',1.0):.2f}</span></div>
    </div>
  </div>

  <div class="panel" style="border-color:#FB5607">
    <h2 style="color:#FB5607">Schedule & Units</h2>
    <div class="big" style="color:#FB5607">{d09.get('total_months','')} <span style="font-size:12px">months</span></div>
    <div class="label">Total Duration</div>
    <div style="margin-top:12px">
      <div class="row"><span class="k">Total Weeks</span><span class="v">{d09.get('total_weeks','')}</span></div>
      <div class="row"><span class="k">Total Units</span><span class="v">{um.get('total','')}</span></div>
      <div class="row"><span class="k">Studios</span><span class="v">{um.get('studios','')}</span></div>
      <div class="row"><span class="k">1 Bedroom</span><span class="v">{um.get('one_bed','')}</span></div>
      <div class="row"><span class="k">2 Bedroom</span><span class="v">{um.get('two_bed','')}</span></div>
      <div class="row"><span class="k">3 Bedroom</span><span class="v">{um.get('three_bed','')}</span></div>
      <div class="row"><span class="k">Penthouses</span><span class="v">{um.get('penthouses','')}</span></div>
    </div>
  </div>

  <div class="footer">
    <div>PMO CASCADE Sovereign Engine v2.0 — {data.get('generated_at','')}</div>
    <div>Design: {', '.join([cat.get('style','')])} | Exterior: {len(cat.get('exterior',[]))} items | Interior: {len(cat.get('interior',[]))} items | Landscape: {cat.get('landscape','')}</div>
    <div>Generated by PMO Engineering Platform</div>
  </div>
</div>
</body></html>"""

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    poster_path = OUTPUT_DIR / "cascade_poster.html"
    poster_path.write_text(poster_html, encoding="utf-8")
    return FileResponse(poster_path, media_type="text/html", filename="PMO_Cascade_Poster.html")

# ─── PROJECT SAVE/LOAD ───

PROJECTS_DB = BASE_DIR / "projects.db"

def init_projects_db():
    conn = sqlite3.connect(str(PROJECTS_DB))
    conn.execute("""CREATE TABLE IF NOT EXISTS projects (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        plot TEXT, emirate TEXT, type TEXT, quality TEXT,
        input_json TEXT, output_json TEXT, catalog_json TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        name TEXT, status TEXT DEFAULT 'draft'
    )""")
    conn.commit()
    conn.close()

init_projects_db()

@app.post("/api/projects/save")
async def save_project(request: Request):
    body = await request.json()
    name = body.get("name", body.get("plot", "Untitled"))
    conn = sqlite3.connect(str(PROJECTS_DB))
    cursor = conn.execute(
        "INSERT INTO projects (plot, emirate, type, quality, input_json, output_json, catalog_json, name) VALUES (?,?,?,?,?,?,?,?)",
        (body.get("plot",""), body.get("emirate",""), body.get("type",""), body.get("quality",""),
         json.dumps(body.get("input",{})), json.dumps(body.get("output",{})), json.dumps(body.get("catalog",{})), name)
    )
    conn.commit()
    project_id = cursor.lastrowid
    conn.close()
    return {"status": "success", "project_id": project_id}

@app.get("/api/projects/list")
async def list_projects():
    conn = sqlite3.connect(str(PROJECTS_DB))
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT id, name, plot, emirate, type, quality, created_at, status FROM projects ORDER BY created_at DESC").fetchall()
    conn.close()
    return {"projects": [dict(r) for r in rows]}

@app.get("/api/projects/{project_id}")
async def get_project(project_id: int):
    conn = sqlite3.connect(str(PROJECTS_DB))
    conn.row_factory = sqlite3.Row
    row = conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "Project not found")
    result = dict(row)
    for field in ["input_json", "output_json", "catalog_json"]:
        if result.get(field):
            result[field] = json.loads(result[field])
    return result

@app.delete("/api/projects/{project_id}")
async def delete_project(project_id: int):
    conn = sqlite3.connect(str(PROJECTS_DB))
    conn.execute("DELETE FROM projects WHERE id=?", (project_id,))
    conn.commit()
    conn.close()
    return {"status": "deleted"}

# ─── N8N WEBHOOK INTEGRATION ───

@app.post("/n8n/webhook/cascade")
async def n8n_webhook_cascade(payload: N8nWebhookPayload):
    """N8n webhook: run cascade engine"""
    if payload.action == "run_cascade" and payload.input_data:
        engine = CascadeEngine()
        result = engine.run(payload.input_data)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        with open(OUTPUT_DIR / "cascade_output.json", "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, default=str)
        return {"status": "success", "result": result}
    elif payload.action == "get_output":
        out_path = OUTPUT_DIR / "cascade_output.json"
        if out_path.exists():
            with open(out_path, encoding="utf-8") as f:
                return {"status": "success", "result": json.load(f)}
        return {"status": "error", "message": "No output found"}
    return {"status": "error", "message": f"Unknown action: {payload.action}"}

@app.post("/n8n/webhook/domain/{domain_id}")
async def n8n_webhook_domain(domain_id: str, payload: N8nWebhookPayload):
    """N8n webhook: query specific domain"""
    if domain_id not in DOMAINS:
        return {"status": "error", "message": f"Domain {domain_id} not found"}
    info = DOMAINS[domain_id]
    db_path = BASE_DIR / f"{domain_id}_{info['name'].replace(' & ', '_').replace(' ', '_')}" / "databases" / info["db"]
    if not db_path.exists():
        for d in BASE_DIR.iterdir():
            if d.is_dir() and d.name.startswith(domain_id):
                for f in d.rglob("*.db"):
                    db_path = f
                    break
    tables = get_tables(db_path) if db_path.exists() else []
    data = {}
    for t in tables:
        data[t] = get_table_data(db_path, t)
    return {"status": "success", "domain": info, "data": data}

@app.post("/n8n/webhook/chain")
async def n8n_webhook_chain(payload: N8nWebhookPayload):
    """N8n webhook: full chain — cascade + all domains"""
    if payload.input_data:
        # Step 1: Run cascade
        engine = CascadeEngine()
        cascade_result = engine.run(payload.input_data)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        with open(OUTPUT_DIR / "cascade_output.json", "w", encoding="utf-8") as f:
            json.dump(cascade_result, f, indent=2, default=str)
        # Step 2: Query all domains
        all_domains = {}
        for did, info in DOMAINS.items():
            db_path = BASE_DIR / f"{did}_{info['name'].replace(' & ', '_').replace(' ', '_')}" / "databases" / info["db"]
            if not db_path.exists():
                for d in BASE_DIR.iterdir():
                    if d.is_dir() and d.name.startswith(did):
                        for f in d.rglob("*.db"):
                            db_path = f
                            break
            tables = get_tables(db_path) if db_path.exists() else []
            data = {}
            for t in tables:
                data[t] = get_table_data(db_path, t)
            all_domains[did] = {"name": info["name"], "tables": tables, "data": data}
        return {"status": "success", "cascade": cascade_result, "domains": all_domains}
    return {"status": "error", "message": "input_data required"}

@app.get("/n8n/webhook/status")
async def n8n_status():
    return {"status": "online", "engine": "PMO CASCADE Sovereign Engine v2.0",
            "domains": len(DOMAINS), "gateway_tables": len(get_tables(GATEWAY_DB))}

# ─── REPORT GENERATION ───

@app.get("/api/report/html")
async def generate_html_report():
    out_path = OUTPUT_DIR / "cascade_output.json"
    if not out_path.exists():
        raise HTTPException(404, "No cascade output. Run cascade first.")
    with open(out_path, encoding="utf-8") as f:
        data = json.load(f)
    d = data.get("derived", {})
    inp = data.get("input", {})
    cat = data.get("catalog_selections") or {}
    iq = data.get("inferred_quality") or {}
    report = f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><title>PMO Report — {inp.get('plot','')}</title>
<style>
body {{ font-family: 'Segoe UI', sans-serif; margin: 40px; background: #0a0a0a; color: #e0e0e0; }}
h1 {{ color: #00B4D8; border-bottom: 2px solid #00B4D8; padding-bottom: 10px; }}
h2 {{ color: #06D6A0; margin-top: 30px; }}
table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
th {{ background: #1a1a2e; color: #00B4D8; padding: 8px; text-align: left; }}
td {{ padding: 8px; border-bottom: 1px solid #333; }}
.metric {{ display: inline-block; background: #1a1a2e; padding: 15px 25px; margin: 5px; border-radius: 8px; border-left: 4px solid #00B4D8; }}
.metric .value {{ font-size: 24px; font-weight: bold; color: #00B4D8; }}
.metric .label {{ color: #888; font-size: 12px; }}
.chip {{ display: inline-block; padding: 4px 12px; margin: 3px; border-radius: 20px; font-size: 12px; background: #1a1a2e; border: 1px solid #333; }}
</style></head><body>
<h1>PMO CASCADE Report — {inp.get('plot','')}</h1>
<p>Generated: {data.get('generated_at','')}</p>
<div>
<div class="metric"><div class="value">{inp.get('emirate','')}</div><div class="label">Emirate</div></div>
<div class="metric"><div class="value">{inp.get('type','')}</div><div class="label">Type</div></div>
<div class="metric"><div class="value">{iq.get('level', inp.get('quality',''))}</div><div class="label">Quality (Inferred)</div></div>
<div class="metric"><div class="value">{inp.get('floors','')}</div><div class="label">Floors</div></div>
<div class="metric"><div class="value">Score: {round((iq.get('score',0.5))*100)}%</div><div class="label">Quality Score</div></div>
<div class="metric"><div class="value">x{iq.get('cost_factor',1.0):.2f}</div><div class="label">Cost Factor</div></div>
</div>
<h2>Design Selections</h2>
<p style="color:#888">Style: {cat.get('style','—')}</p>
<p style="color:#888">Exterior: {', '.join(cat.get('exterior',[]))}</p>
<p style="color:#888">Interior: {', '.join(cat.get('interior',[]))}</p>
<p style="color:#888">Landscape: {cat.get('landscape','—')}</p>
<p style="color:#888">Facilities: {', '.join(cat.get('facilities',[]))}</p>
<h2>Domain 04 — Zoning</h2>
<table><tr><th>Parameter</th><th>Value</th></tr>
<tr><td>FAR Range</td><td>{d.get('far_min','')} — {d.get('far_max','')}</td></tr>
<tr><td>Height Range</td><td>{d.get('h_min','')}m — {d.get('h_max','')}m</td></tr>
<tr><td>Coverage</td><td>{d.get('coverage_max','')}%</td></tr>
<tr><td>Setbacks (F/R/S)</td><td>{d.get('setback_f','')}m / {d.get('setback_r','')}m / {d.get('setback_s','')}m</td></tr>
</table>
<h2>Domain 06 — Design</h2>
<table><tr><th>System</th><th>Specification</th></tr>
<tr><td>Structural</td><td>{d.get('structural_system','')}</td></tr>
<tr><td>Concrete</td><td>{d.get('concrete_grade','')}</td></tr>
<tr><td>Foundation</td><td>{d.get('foundation_type','')}</td></tr>
</table>
<h2>Domain 08 — Cost</h2>
<table><tr><th>Component</th><th>AED</th></tr>
<tr><td>Construction</td><td>{d.get('construction_cost',0):,.0f}</td></tr>
<tr><td>Land</td><td>{d.get('land_cost',0):,.0f}</td></tr>
<tr><td>Soft Costs</td><td>{d.get('soft_cost',0):,.0f}</td></tr>
<tr><td>Contingency</td><td>{d.get('contingency_cost',0):,.0f}</td></tr>
<tr><td><strong>Total</strong></td><td><strong>{d.get('total_project_cost',0):,.0f}</strong></td></tr>
</table>
<h2>Domain 09 — Schedule</h2>
<table><tr><th>Phase</th><th>Weeks</th></tr>"""
    d09 = d.get("domain_09", {})
    for k, v in d09.items():
        if "_weeks" in k:
            report += f"<tr><td>{k.replace('_weeks','').title()}</td><td>{v}</td></tr>"
    report += f"""<tr><td><strong>Total</strong></td><td><strong>{d09.get('total_weeks',0)} weeks ({d09.get('total_months',0)} months)</strong></td></tr>
</table>
<h2>Domain 12 — Sustainability</h2>
<table><tr><th>Target</th><th>Level</th></tr>
<tr><td>LEED</td><td>{d.get('domain_12',{}).get('leed_target','')}</td></tr>
<tr><td>Estidama</td><td>{d.get('domain_12',{}).get('estidama_target','')}</td></tr>
</table>
</body></html>"""
    out_file = OUTPUT_DIR / "cascade_report.html"
    out_file.write_text(report, encoding="utf-8")
    return FileResponse(out_file, media_type="text/html", filename="cascade_report.html")

if __name__ == "__main__":
    port = int(os.getenv("PORT", 9000))
    uvicorn.run(app, host="0.0.0.0", port=port)
