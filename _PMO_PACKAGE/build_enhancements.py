#!/usr/bin/env python3
"""
PMO FIX & ENHANCE — Fix broken formulas + Add hyperlinks + Enhance workbooks
"""
import sqlite3, csv, os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
LINK_FONT = Font(name="Calibri", size=10, color="2980B9", underline="single")
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ============================================================================
# FIX BROKEN FORMULAS IN ORIGINAL FILES
# ============================================================================
def fix_original_formulas():
    print("\n=== FIXING BROKEN FORMULAS IN ORIGINAL FILES ===")
    
    # Fix Feasibility.csv — the formula =D4-C4-B4 is off by one row
    csv_dir = BASE.parent / "the_Full_Set/full-session/_nested_extracted/DesigningaMatrixforFloorTypeandUnitOptimization"
    feasibility_file = csv_dir / "Feasibility.csv"
    if feasibility_file.exists():
        lines = feasibility_file.read_text().split("\n")
        fixed_lines = [lines[0]]  # header
        for i in range(1, len(lines)):
            if lines[i].strip():
                parts = lines[i].split(",")
                if len(parts) >= 5:
                    # Fix: Net Cash Flow = Revenue - Opex - Capex
                    # Original had wrong row references
                    parts[4] = f"=D{i+1}-C{i+1}-B{i+1}"
                    lines[i] = ",".join(parts)
            fixed_lines.append(lines[i])
        feasibility_file.write_text("\n".join(fixed_lines))
        print("  [FIXED] Feasibility.csv — Net Cash Flow formulas corrected")

    # Fix CostControl.csv — Variance formulas
    cost_file = csv_dir / "CostControl.csv"
    if cost_file.exists():
        lines = cost_file.read_text().split("\n")
        fixed_lines = [lines[0]]  # header
        for i in range(1, len(lines)):
            if lines[i].strip():
                parts = lines[i].split(",")
                if len(parts) >= 7:
                    # Fix: Variance (AED) = Budget - Actual
                    parts[5] = f"=C{i+1}-E{i+1}"
                    # Fix: Variance % = IF(Budget=0,0,Variance/Budget)
                    parts[6] = f'=IF(C{i+1}=0,0,F{i+1}/C{i+1})'
                    lines[i] = ",".join(parts)
            fixed_lines.append(lines[i])
        cost_file.write_text("\n".join(fixed_lines))
        print("  [FIXED] CostControl.csv — Variance formulas corrected")

    # Fix Construction.csv — Variance % formula
    constr_file = csv_dir / "Construction.csv"
    if constr_file.exists():
        lines = constr_file.read_text().split("\n")
        fixed_lines = [lines[0]]  # header
        for i in range(1, len(lines)):
            if lines[i].strip():
                parts = lines[i].split(",")
                if len(parts) >= 7:
                    # Fix: Variance % = Actual% - Planned%
                    parts[5] = f"=E{i+1}-D{i+1}"
                    lines[i] = ",".join(parts)
            fixed_lines.append(lines[i])
        constr_file.write_text("\n".join(fixed_lines))
        print("  [FIXED] Construction.csv — Variance % formulas corrected")

# ============================================================================
# ADD HYPERLINKS TO WORKBOOKS
# ============================================================================
def add_hyperlinks_to_workbooks():
    print("\n=== ADDING HYPERLINKS TO WORKBOOKS ===")
    
    domains = [
        ("01_PROJECT_AND_PLOT", "project_plot.db", "01_project_plot.sql"),
        ("02_LOCATION_GIS", "location_gis.db", "02_location_gis.sql"),
        ("03_STAKEHOLDERS", "stakeholders.db", "03_stakeholders.sql"),
        ("04_ZONING_REGULATORY", "zoning_regulatory.db", "04_zoning_regulatory.sql"),
        ("05_VALIDATION_COMPLIANCE", "validation.db", "05_validation.sql"),
        ("06_DESIGN_PARAMETERS", "design_parameters.db", "06_design.sql"),
        ("07_UNIT_MIX_PROGRAM", "unit_mix.db", "07_unit_mix.sql"),
        ("08_COST_ECONOMICS", "cost_economics.db", "08_cost_economics.sql"),
        ("09_SCHEDULE_TIMELINE", "schedule.db", "09_schedule.sql"),
        ("10_APPROVALS_AUTHORITIES", "approvals.db", "10_approvals.sql"),
        ("11_GEOTECHNICAL", "geotechnical.db", "11_geotechnical.sql"),
        ("12_SUSTAINABILITY_QUALITY_BIM", "sustainability.db", "12_sustainability.sql"),
    ]
    
    for domain_name, db_file, sql_file in domains:
        xlsx_path = BASE / domain_name / "workbook.xlsx"
        if not xlsx_path.exists():
            continue
        
        try:
            wb = load_workbook(str(xlsx_path))
            
            # Add a Hyperlinks sheet
            if "Links" not in wb.sheetnames:
                ws = wb.create_sheet("Links")
                ws.sheet_properties.tabColor = "2980B9"
                
                # Title
                ws.cell(row=1, column=1, value="Quick Links").font = Font(name="Calibri", bold=True, color="2C3E50", size=14)
                
                # Headers
                for c, h in enumerate(["File Type", "File Name", "Path"], 1):
                    cell = ws.cell(row=3, column=c, value=h)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                
                # Links
                links = [
                    ("Database", db_file, str(BASE / domain_name / "databases" / db_file)),
                    ("Schema", sql_file, str(BASE / domain_name / "schemas" / sql_file)),
                    ("CSV Data", "data_raw/", str(BASE / domain_name / "data_raw")),
                    ("Diagrams", "diagrams/", str(BASE / domain_name / "diagrams")),
                ]
                
                for r, (ftype, fname, fpath) in enumerate(links, 4):
                    ws.cell(row=r, column=1, value=ftype).font = BOLD_FONT
                    ws.cell(row=r, column=1).border = THIN_BORDER
                    ws.cell(row=r, column=2, value=fname).font = LINK_FONT
                    ws.cell(row=r, column=2).border = THIN_BORDER
                    if os.path.exists(fpath) or os.path.isdir(fpath):
                        ws.cell(row=r, column=2).hyperlink = fpath
                    ws.cell(row=r, column=3, value=fpath).font = Font(name="Consolas", size=9, color="7F8C8D")
                    ws.cell(row=r, column=3).border = THIN_BORDER
                
                ws.column_dimensions["A"].width = 15
                ws.column_dimensions["B"].width = 30
                ws.column_dimensions["C"].width = 80
            
            # Add DB summary to first sheet
            ws_first = wb.active
            summary_row = ws_first.max_row + 3
            ws_first.cell(row=summary_row, column=1, value="Database Summary").font = Font(name="Calibri", bold=True, color="2C3E50", size=12)
            
            db_path = BASE / domain_name / "databases" / db_file
            if db_path.exists():
                conn = sqlite3.connect(str(db_path))
                cur = conn.cursor()
                cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
                tables = [r[0] for r in cur.fetchall()]
                
                for t_idx, table in enumerate(tables):
                    r = summary_row + 1 + t_idx
                    cur.execute(f"SELECT COUNT(*) FROM [{table}]")
                    count = cur.fetchone()[0]
                    ws_first.cell(row=r, column=1, value=f"Table: {table}").font = BOLD_FONT
                    ws_first.cell(row=r, column=1).border = THIN_BORDER
                    ws_first.cell(row=r, column=2, value=f"{count} rows").font = Font(name="Calibri", size=10, color="27AE60" if count > 0 else "E74C3C")
                    ws_first.cell(row=r, column=2).border = THIN_BORDER
                
                conn.close()
            
            wb.save(str(xlsx_path))
            print(f"  [LINKS] {domain_name}/workbook.xlsx — hyperlinks added")
        except Exception as e:
            print(f"  [ERR] {domain_name}: {e}")

# ============================================================================
# ENHANCE MASTER INDEX WITH CROSS-LINKS
# ============================================================================
def enhance_master_index():
    print("\n=== ENHANCING MASTER_INDEX.xlsx ===")
    
    xlsx_path = BASE / "MASTER_INDEX.xlsx"
    if not xlsx_path.exists():
        return
    
    wb = load_workbook(str(xlsx_path))
    ws = wb.active
    
    # Find the data rows and add hyperlinks
    for row in range(6, 18):  # Data rows
        domain_id = ws.cell(row=row, column=1).value
        if domain_id and str(domain_id).strip().isdigit():
            domain_dir = BASE / f"{domain_id.zfill(2)}_*"
            # Find actual directory
            for d in BASE.iterdir():
                if d.is_dir() and d.name.startswith(f"{domain_id.zfill(2)}"):
                    xlsx_file = d / "workbook.xlsx"
                    if xlsx_file.exists():
                        ws.cell(row=row, column=2).hyperlink = str(xlsx_file)
                        ws.cell(row=row, column=2).font = Font(name="Calibri", size=10, color="2980B9", underline="single")
                    break
    
    wb.save(str(xlsx_path))
    print("  [DONE] MASTER_INDEX.xlsx enhanced with cross-links")

# ============================================================================
# CREATE HTML DASHBOARD
# ============================================================================
def create_dashboard():
    print("\n=== CREATING HTML DASHBOARD ===")
    
    html = """<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PMO DOMAINS — لوحة المعلومات</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Segoe UI', Tahoma, sans-serif; background: #1a1a2e; color: #eee; }
        .header { background: linear-gradient(135deg, #2C3E50, #3498DB); padding: 30px; text-align: center; }
        .header h1 { font-size: 28px; color: #fff; margin-bottom: 10px; }
        .header p { color: #BDC3C7; font-size: 14px; }
        .grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(350px, 1fr)); gap: 20px; padding: 30px; }
        .card { background: #16213e; border-radius: 12px; padding: 20px; border-left: 5px solid; transition: transform 0.2s; }
        .card:hover { transform: translateY(-5px); }
        .card h3 { font-size: 16px; margin-bottom: 10px; }
        .card .ar { color: #BDC3C7; font-size: 13px; margin-bottom: 15px; }
        .card .stats { display: flex; gap: 15px; margin-bottom: 15px; }
        .stat { text-align: center; }
        .stat .num { font-size: 24px; font-weight: bold; }
        .stat .label { font-size: 11px; color: #7F8C8D; }
        .card .files { font-size: 12px; color: #95A5A6; }
        .card .files span { color: #3498DB; }
        .summary { background: #0f3460; padding: 20px; margin: 0 30px 30px; border-radius: 12px; display: flex; justify-content: space-around; text-align: center; }
        .summary .item .num { font-size: 32px; font-weight: bold; color: #3498DB; }
        .summary .item .label { font-size: 12px; color: #BDC3C7; }
    </style>
</head>
<body>
    <div class="header">
        <h1>PMO DOMAINS — نظام إدارة المشاريع</h1>
        <p>12 مجالاً | 69 ملفاً | بيانات هندسية مُضيّقة | Emirates: Dubai, Abu Dhabi, Sharjah, RAK, Ajman, UAQ, Fujairah</p>
    </div>
    
    <div class="summary">
        <div class="item"><div class="num">12</div><div class="label">مجال</div></div>
        <div class="item"><div class="num">12</div><div class="label">Schema SQL</div></div>
        <div class="item"><div class="num">17</div><div class="label">CSV Data</div></div>
        <div class="item"><div class="num">13</div><div class="label">Database DB</div></div>
        <div class="item"><div class="num">12</div><div class="label">Diagram</div></div>
        <div class="item"><div class="num">13</div><div class="label">Workbook XLSX</div></div>
    </div>

    <div class="grid">
"""
    
    domains = [
        ("01", "Project & Plot", "المشروع والأراضي", "00B4D8", "5", "1", "1", "1", "project_data.csv"),
        ("02", "Location & GIS", "الموقع والخرائط", "06D6A0", "10", "1", "1", "1", "location_data.csv"),
        ("03", "Stakeholders", "أصحاب المصلحة", "FFD166", "5", "1", "1", "1", "stakeholders_data.csv"),
        ("04", "Zoning & Regulatory", "التخطيط والتنظيم", "EF476F", "7", "1", "1", "1", "far_limits.csv"),
        ("05", "Validation & Compliance", "المطابقة والتوثيق", "118AB2", "8", "2", "1", "1", "compliance_data.csv + enum_lookups.csv"),
        ("06", "Design Parameters", "التصميم والمعايير", "073B4C", "9", "2", "1", "1", "cost_indices.csv + platform_features.csv"),
        ("07", "Unit Mix & Program", "الوحدات والبرامج", "8338EC", "10", "1", "1", "1", "unit_types.csv"),
        ("08", "Cost & Economics", "التكاليف والتحليل المالي", "FF006E", "15", "3", "1", "1", "csi_cost_codes.csv + cost_benchmarks.csv + deliverables_pricing.csv"),
        ("09", "Schedule & Timeline", "الخطة الزمنية", "FB5607", "23", "1", "1", "1", "wbs_activities.csv"),
        ("10", "Approvals & Authorities", "الموافقات والسلطات", "3A86FF", "14", "2", "1", "1", "permit_sequence.csv + authority_matrix_raw.csv"),
        ("11", "Geotechnical", "الجيوقيقة والأساسيات", "80ED99", "26", "3", "1", "1", "soil_types.csv + site_geotechnical.csv + soil_reference_raw.csv"),
        ("12", "Sustainability & BIM", "الاستدامة والجودة", "C77DFF", "10", "2", "1", "1", "risk_catalog.csv + risks_catalog_raw.csv"),
    ]
    
    for did, name, name_ar, color, rows, csvs, db, diag, desc in domains:
        html += f"""
        <div class="card" style="border-color: #{color}">
            <h3 style="color: #{color}">{did} — {name}</h3>
            <div class="ar">{name_ar}</div>
            <div class="stats">
                <div class="stat"><div class="num" style="color: #{color}">{rows}</div><div class="label">صف</div></div>
                <div class="stat"><div class="num">{csvs}</div><div class="label">CSV</div></div>
                <div class="stat"><div class="num">{db}</div><div class="label">DB</div></div>
                <div class="stat"><div class="num">{diag}</div><div class="label">Diagram</div></div>
            </div>
            <div class="files">Files: <span>{desc}</span></div>
        </div>
"""
    
    html += """
    </div>
    
    <div style="background: #0f3460; padding: 20px; margin: 0 30px 30px; border-radius: 12px;">
        <h3 style="color: #3498DB; margin-bottom: 15px;">النطاقات المُضيّقة</h3>
        <table style="width: 100%; border-collapse: collapse; color: #BDC3C7;">
            <tr style="border-bottom: 1px solid #2C3E50;">
                <th style="padding: 8px; text-align: right;">النطاق الأصلي</th>
                <th style="padding: 8px; text-align: right;">القيمة المضيّقة</th>
                <th style="padding: 8px; text-align: right;">المصدر</th>
            </tr>
            <tr style="border-bottom: 1px solid #1a1a2e;">
                <td style="padding: 8px;">1-2 basements</td>
                <td style="padding: 8px; color: #27AE60;">min:1, max:2, typical:2</td>
                <td style="padding: 8px;">T_GEOTECH</td>
            </tr>
            <tr style="border-bottom: 1px solid #1a1a2e;">
                <td style="padding: 8px;">7-14 days (Land Title)</td>
                <td style="padding: 8px; color: #27AE60;">min:7, max:14, avg:10</td>
                <td style="padding: 8px;">MV1_Permit</td>
            </tr>
            <tr style="border-bottom: 1px solid #1a1a2e;">
                <td style="padding: 8px;">30-60 days (Building Permit)</td>
                <td style="padding: 8px; color: #27AE60;">min:30, max:60, avg:45</td>
                <td style="padding: 8px;">MV1_Permit</td>
            </tr>
            <tr style="border-bottom: 1px solid #1a1a2e;">
                <td style="padding: 8px;">200-400 kPa (Dense Sand)</td>
                <td style="padding: 8px; color: #27AE60;">min:200, max:400, typical:300</td>
                <td style="padding: 8px;">MV1_Soil</td>
            </tr>
            <tr style="border-bottom: 1px solid #1a1a2e;">
                <td style="padding: 8px;">15-20% yield</td>
                <td style="padding: 8px; color: #27AE60;">min:15, max:20, typical:17.5</td>
                <td style="padding: 8px;">RM Model</td>
            </tr>
        </table>
    </div>

    <div style="text-align: center; padding: 20px; color: #7F8C8D; font-size: 12px;">
        Generated: 2026-06-16 | Style: Ras Flat Dark | Font: Calibri Bold
    </div>
</body>
</html>"""
    
    dashboard_path = BASE / "DASHBOARD.html"
    dashboard_path.write_text(html, encoding="utf-8")
    print(f"  [DONE] DASHBOARD.html created ({len(html)} bytes)")

# ============================================================================
# MAIN
# ============================================================================
if __name__ == "__main__":
    print("=" * 60)
    print("PMO FIX & ENHANCE — Formulas + Hyperlinks + Dashboard")
    print("=" * 60)
    fix_original_formulas()
    add_hyperlinks_to_workbooks()
    enhance_master_index()
    create_dashboard()
    print("=" * 60)
    print("ALL ENHANCEMENTS COMPLETE")
    print("=" * 60)
